import io
import base64
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from services.file_rule_maker import FileRuleMaker
from services.file_validator import FileValidator
from services.merge_excel_exporter import MergeExcelExporter
from services.split_excel_exporter import SplitExcelExporter
from io import BytesIO
from utils import excel_processor
import json
import openpyxl as px
import logging
from werkzeug.utils import secure_filename
import utils.excel_processor as XPRO
import copy

app = Flask(__name__)
CORS(app)
app.logger.setLevel(logging.INFO)

fuker = FileRuleMaker()
fileValidator = FileValidator()
mergeExcelExporter = MergeExcelExporter()


@app.route("/save_rawFile", methods=["POST"])
def save_raw_file():

    file = request.files.get("file")
    file_name = secure_filename(file.filename)
    file_stream = io.BytesIO(file.read())

    if file:
        # 转换处理

        if file_name.endswith(".xls"):
            file_stream = excel_processor.Excel_IO().convert_excel_format(
                file_stream, "xls", "xlsx", True
            )
            # file_stream.seek(0)
            # print("file_stream:", file_stream)

        # 给file_rule_maker 的属性赋值
        fuker.get_file_stream(file_stream, file_name)

        # 发送处理后的文件给前端
        byte_stream = io.BytesIO()
        byte_stream.write(file_stream.getvalue())
        byte_stream.seek(0)  # 跳转到流的开头
        return send_file(
            byte_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=file_name,
        )


@app.route("/generate_user_rule_dict", methods=["POST"])
def generate_user_rule_dict():
    fields_index_col_list = json.loads(request.form.get("fields"))
    fields_index_col_dict = {
        field["position"]: field["fieldName"] for field in fields_index_col_list
    }
    # print(file, file_name,fields_index_col_dict)

    if fields_index_col_dict:
        # print("fields_index_col_dict: ", fields_index_col_dict)
        field_rules = fuker.generate_user_rule_dict(fields_index_col_dict)
        # print("field rules: ", field_rules)
        return jsonify(field_rules)
    return "Error processing file", 500


@app.route("/create_final_rules_and_examples", methods=["POST"])
def create_final_rules_and_examples_file():
    selected_field_rules = json.loads(request.form.get("finalRules"))
    # print("selected_field_rules: ", selected_field_rules)
    final_rules_and_examples, simulate_rule_excel_stream_dict = (
        fuker.create_final_rules_and_examples(selected_field_rules)
    )

    # 发送处理后的文件给前端
    file_data = {}
    for mode, simulate_rule_excel_stream in simulate_rule_excel_stream_dict.items():
        byte_stream = io.BytesIO()
        byte_stream.write(simulate_rule_excel_stream.getvalue())
        byte_stream.seek(0)  # 跳转到流的开头

        # 将数据流转换为Base64编码的字符串
        file_data[mode] = [
            base64.b64encode(byte_stream.getvalue()).decode("utf-8"),
            final_rules_and_examples[mode],
        ]

    # print("file_data", file_data)
    return jsonify(file_data), 200


@app.route("/load_and_check_data", methods=["POST"])
def load_and_check_data():
    excelFile = request.files.get("excelFile")
    excel_stream = io.BytesIO(excelFile.read())
    excel_stream.seek(0)
    excelFile_name = secure_filename(excelFile.filename)

    if excelFile_name.endswith(".xls"):
        excel_stream = excel_processor.Excel_IO().convert_excel_format(
            excel_stream, "xls", "xlsx", True
        )

    ruleFile = request.files.get("ruleFile")

    rule_dict = json.loads(ruleFile.read().decode("utf-8"))
    # print("rule_dict:", rule_dict)

    new_excel = fileValidator.get_files_stream(excel_stream, excelFile_name, rule_dict)
    # print(new_excel)
    error_cell_info, checked_excel = fileValidator.validate_filled_excel(new_excel)

    # XPRO.Excel_IO().load_workbook_from_stream(checked_excel)[0].save("ddd.xlsx")

    # print("range_and_rule", range_and_rule)

    # print("error_index_col", error_index_col)
    checked_excel_error = {
        "error_cell_info": error_cell_info,
        "checked_excel": base64.b64encode(checked_excel.getvalue()).decode("utf-8"),
    }

    # 发送处理后的文件给前端
    return jsonify(checked_excel_error), 200


@app.route("/check_data", methods=["POST"])
def check_data():
    # print("request.files",request.files)
    excelFile = request.files.get("excelFile")
    excel_stream = io.BytesIO(excelFile.read())
    error_cell_info, checked_excel = fileValidator.validate_filled_excel(excel_stream)
    # print("error_index_col:", error_index_col)
    excel_stream.seek(0)

    # 发送处理后的文件给前端
    checked_excel_error = {
        "error_cell_info": error_cell_info,
        "checked_excel": base64.b64encode(checked_excel.getvalue()).decode("utf-8"),
    }

    # 发送处理后的文件给前端
    return jsonify(checked_excel_error), 200


# 处理上传的文件群和样表文件
@app.route("/load-excelFiles-example", methods=["POST"])
def load_excelFiles_example():
    excelFiles = request.files.getlist("excelFiles")
    exampleFile = request.files.get("exampleFile")

    excel_stream_dict = {}

    for excelFile in excelFiles:
        excel_stream = io.BytesIO(excelFile.read())
        excel_filename = secure_filename(excelFile.filename)
        if excel_filename.endswith(".xls"):
            excel_stream = excel_processor.Excel_IO().convert_excel_format(
                excel_stream, "xls", "xlsx", True
            )
        excel_stream.seek(0)
        excel_stream_dict[excelFile.filename] = excel_stream

    mergeExcelExporter.acquire_excel_group(excel_stream_dict)

    example_stream = io.BytesIO(exampleFile.read())
    example_filename = secure_filename(exampleFile.filename)

    if example_filename.endswith(".xls"):
        example_stream = excel_processor.Excel_IO().convert_excel_format(
            example_stream, "xls", "xlsx", True
        )
    # example_stream.seek(0)
    byte_stream = io.BytesIO()
    byte_stream.write(example_stream.getvalue())
    byte_stream.seek(0)  # 跳转到流的开头
    # 发送处理后的文件给前端
    return send_file(
        byte_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=example_filename,
    )


@app.route("/extract_example_info", methods=["POST"])
def extract_example_info():
    # 获取文件
    excel_blob = request.files.get("excelBlob")
    # 获取 ranges 和 startRow，因为它们以文本形式发送，所以使用 request.form
    ranges = request.form.get("ranges")
    start_row = request.form.get("startRow")

    print("ranges", ranges)
    print("start_row", start_row)
    excel_stream = io.BytesIO(excel_blob.read())

    mergeExcelExporter.extract_template_info(
        header_range=ranges, data_start_row=start_row, template_excel=excel_stream
    )

    recheck_excel_info = {
        "check_info": "",
        "recheck_excel": "",
        "recheck_excel_fileName": "",
        "checked_count": 0,
    }

    for file_name in copy.deepcopy(
        list(mergeExcelExporter.original_excel_group.keys())
    ):
        [check_wb, check_ws] = mergeExcelExporter.original_excel_group[file_name]
        # recheck_wb,recheck_ws=Excel_IO.load_workbook_from_stream(needed_check_excel_stream)
        check_flag, check_info = mergeExcelExporter.verify_excel(check_wb, check_ws)
        mergeExcelExporter.update_excel_files(check_flag, file_name)
        print(
            "len(mergeExcelExporter.verified_excel_group)",
            len(mergeExcelExporter.verified_excel_group),
        )

        if not check_flag:
            recheck_excel_stream = XPRO.Excel_IO().stream_excel_to_frontend(check_wb)

            recheck_excel_info["check_info"] = check_info
            recheck_excel_info["recheck_excel"] = base64.b64encode(
                recheck_excel_stream.getvalue()
            ).decode("utf-8")

            recheck_excel_info["recheck_excel_fileName"] = file_name
            break
    print(
        "len(mergeExcelExporter.verified_excel_group)",
        len(mergeExcelExporter.verified_excel_group),
    )
    recheck_excel_info["checked_count"] = len(mergeExcelExporter.verified_excel_group)

    return jsonify(recheck_excel_info), 200


@app.route("/recheck", methods=["POST"])
def recheck():

    recheck_blob = request.files.get("recheckExcelBlob")
    recheck_stream = io.BytesIO(recheck_blob.read())

    print(request.form.get("recheck_fileName"))
    recheck_fileName = request.form.get("recheck_fileName")

    recheck_wb, recheck_ws = XPRO.Excel_IO().load_workbook_from_stream(recheck_stream)

    recheck_stream.seek(0)

    recheck_excel_info = {
        "check_info": "",
        "recheck_excel": "",
        "recheck_excel_fileName": "",
        "checked_count": len(mergeExcelExporter.verified_excel_group),
    }

    recheck_flag, recheck_info = mergeExcelExporter.verify_excel(recheck_wb, recheck_ws)

    if not recheck_flag:
        recheck_excel_info["check_info"] = recheck_info
        recheck_excel_info["recheck_excel"] = base64.b64encode(
            recheck_stream.getvalue()
        ).decode("utf-8")
        recheck_excel_info["recheck_excel_fileName"] = recheck_fileName

    else:
        mergeExcelExporter.verified_excel_group[recheck_fileName] = [
            recheck_wb,
            recheck_ws,
        ]
        for file_name in copy.deepcopy(
            list(mergeExcelExporter.original_excel_group.keys())
        ):
            [check_wb, check_ws] = mergeExcelExporter.original_excel_group[file_name]
            # recheck_wb,recheck_ws=Excel_IO.load_workbook_from_stream(needed_check_excel_stream)
            check_flag, check_info = mergeExcelExporter.verify_excel(check_wb, check_ws)
            mergeExcelExporter.update_excel_files(check_flag, file_name)
            if not check_flag:
                recheck_excel_stream = XPRO.Excel_IO().stream_excel_to_frontend(
                    check_wb
                )

                recheck_excel_info["check_info"] = check_info
                recheck_excel_info["recheck_excel"] = (
                    base64.b64encode(recheck_excel_stream.getvalue()).decode("utf-8"),
                )

                recheck_excel_info["recheck_excel_fileName"] = file_name
                break

    recheck_excel_info["checked_count"] = len(mergeExcelExporter.verified_excel_group)
    print(
        "len(mergeExcelExporter.verified_excel_group)",
        len(mergeExcelExporter.verified_excel_group),
    )

    return jsonify(recheck_excel_info), 200


@app.route("/contact", methods=["POST"])
def contact():
    from os.path import join as J

    print(len(mergeExcelExporter.verified_excel_group))
    contacted_stream = mergeExcelExporter.merge_and_format_excels()
    # XPRO.Excel_IO().load_workbook_from_stream(contacted_stream)[0].save(
    #     J(r"tests\for_concat\for_func5_merge_and_format_excels", "总表.xlsx")
    # )

    byte_stream = io.BytesIO()
    byte_stream.write(contacted_stream.getvalue())
    byte_stream.seek(0)  # 跳转到流的开头
    # 发送处理后的文件给前端
    return send_file(
        byte_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Contacted.xlsx",
    )


@app.route("/load_split_parameters", methods=["POST"])
def load_split_parameters():
    splitExcelExporter = SplitExcelExporter()

    summaryExcelBlob = request.files.get("summaryExcelBlob")
    referenceColumn = request.form.get("referenceColumn")
    startRow = request.form.get("startRow")

    summaryExcel_stream = io.BytesIO(summaryExcelBlob.read())

    splitExcelExporter.load_excel_parameters(
        summaryExcel_stream, int(startRow), referenceColumn
    )

    split_dict = splitExcelExporter.split_worksheet()
    print("split_dict", split_dict)
    for k, v in split_dict.items():
        if k == None:
            split_dict[k] = "缺失值" + ".xlsx"
        split_dict[k] = k + ".xlsx"

    zip_stream = splitExcelExporter.zip_split_files(split_dict)

    return send_file(
        zip_stream,
        mimetype="application/zip",
        as_attachment=True,
        download_name="拆分后.zip",
    )


if __name__ == "__main__":
    app.run(debug=True, port=1129)
