import os, sys, io

# import pandas as pd
import openpyxl as px
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.utils import (
    get_column_letter,
    coordinate_to_tuple,
    range_boundaries,
    column_index_from_string,
)
from typing import IO, List, Dict, Union

"""用于导入项目中不在同一文件夹的库"""
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import utils.excel_processor as XPRO
from utils.string_processor import *


class MergeExcelExporter:
    def __init__(self):
        """初始化方法，创建一个DataFrame和excel"""
        self.data_start_row, self.data_row_height = None, None  #
        self.original_excel_group = {}  # acquire_excel_group
        self.verified_excel_group = {}  # verify_excel
        self.each_file_max_row_width = (
            {}
        )  # 有待商榷每列在各个文件的最大宽度 进一步：模式：完全样表模式和最大模式
        self.final_excel_wb, self.final_excel_ws, self.template_info = None, None, {}
        self.header_row = []  # extract_template_info
        self.needed_col = []  # extract_template_info
        self.have_seprows = False  # extract_template_info
        self.not_needed_col = (
            []
        )  ##extract_template_info进一步。最终删掉不需要的这么个删法
        self.header_max_col, self.header_max_row = 0, 0  # extract_template_info
        self.row_after_header = None  ##extract_template_info
        self.template_header_content = None  ##extract_template_info
        self.template_data_style = (
            {}
        )  ##extract_template_info中update，key为column letter
        self.Xio = XPRO.Excel_IO()  # 自动创建，读写全部用这个对象读取。
        self.Xattr = None  # 以样表为参数，extract_template_info
        self.group_datanum_dict = {}  # merge

    def acquire_excel_group(self, stream_data: dict) -> list:
        """
        从数据流获取：用于合并的excel文件群
        输出到数据流：非Excel文件的文件名列表

        Parameters from stream:
            stream_data (io.BytesIO in dict):
                content: 包含文件名和文件内容的字典
                format : {"file_name1":io.BytesIO, "file_name2":io.BytesIO}
        Returns to stream:
            non_excel_file_names (list):
                content: 所有非Excel文件的文件名列表
                format : ["non_xfile_name1","non_xfile_name2","non_xfile_name3"]
        """
        # 获取文件名，若有非excel文件，直接retun
        file_names = [
            non_xfile_name
            for non_xfile_name in stream_data
            if not non_xfile_name.endswith((".xlsx", ".xls"))
        ]
        if file_names:
            return file_names

        self.original_excel_group = {
            merge_excel_filename: list(
                self.Xio.load_workbook_from_stream(merge_excel_stream)
            )
            for merge_excel_filename, merge_excel_stream in stream_data.items()
        }
        return []

    def extract_template_info(
        self,
        header_range: str,
        data_start_row: Union[str, int],
        template_excel: IO[bytes],
    ):
        r"""
        从数据流获取：表头的范围；字段行；数据开始行；样表名与样表文件的字典
        功能：从样表中提取表头的内容、样式、位置；数据的位置、样式；样表
        输出到数据流：None

        Parameters from stream:
            header_range (cell_ranges_str):# 表头中可以个别列不要。
                content: string中以,间隔的的多个彼此不连续、不交叉的单元格区域，各个单元格区域内部连续
                each_range_pattern:RANGE_EXPR = r'''
                                        [$]?(?P<min_col>[A-Za-z]{1,3})?
                                        [$]?(?P<min_row>\d+)?
                                        (:[$]?(?P<max_col>[A-Za-z]{1,3})?
                                        [$]?(?P<max_row>\d+)?)?
                                        '''# 转行是为了更好地看各部分情况，实际不必转行
                format : "A1:B2,A3,$D1:$D2,$E:$E,3:4,5:5"
            data_start_row (str):#进一步，前端设置默认值为表头下一行，用户进行修改，提示填写此内容是因为有时数据前有一行不需要；另外，数据结束行也很重要
                content: 数据开始行号的string
                format : "3"
            template_excel (excel_stream): 用户选择的作为样表的Excel文件，前端已自行转化为xlsx(若需)，用户已根据需求将其格式进行修改(若需)

        Returns to stream:
            template_info (dict):
                content: None #样表的表头、数据位置、内容、样式信息

        #   参数field_row: str先不考虑；进一步：确定一下字段行，用于查看字段列是否颠倒？,
        # field_row(str):#进一步，前端设置默认值为表头最后一行，用户进行修改，提示填写此内容是因为有时各列数据位置发生交换
        #             content: 数据开始行号的string
        #             format : "3"
        """
        # 读取样表,直接在样表上做处理
        self.final_excel_wb, self.final_excel_ws = self.Xio.load_workbook_from_stream(
            template_excel
        )
        self.Xattr = XPRO.Excel_attribute(self.final_excel_wb, self.final_excel_ws)
        self.data_start_row = int(data_start_row)

        # TODO：
        # 从样表中提取表头的内容、样式、位置
        # 数据的位置、样式；样表
        # 编写获取区域样式的方法，编写获取区域值的方法，获取表头的样式、内容，赋给样表ws作为最终合并ws

        needed_col = (
            set()
        )  # 根据字段列推数据列，合并时全部列合并，最后再统一删除列。数据样式也只给这些列。进一步：列颠倒问题。
        header_row = set()
        for one_range in header_range.split(","):
            min_col, min_row, max_col, max_row = range_boundaries(one_range)
            needed_col.update(list(range(min_col, max_col + 1)))
            header_row.update(list(range(min_row, max_row + 1)))
            # 整行整列的情况，默认px识别的最大行列
            max_col = self.final_excel_ws.max_column if not max_col else max_col
            max_row = self.final_excel_ws.max_row if not max_row else max_row
            # self.template_header_content.update(self.Xattr.get_range_cells_dict(min_col, min_row, max_col, max_row))
            # 不是字典而是df了

        # 先读取数据起始行的除了超链接以外的所有属性，然后去掉选字段时未选入的列
        t_d_s = self.Xattr.get_row_attributes(data_start_row, ["hyperlink"])
        self.template_data_style.update(
            {
                coordinate_from_string(cell)[0]: cell_attr
                for cell, cell_attr in t_d_s.items()
                if range_boundaries(cell)[0] in needed_col
            }
        )
        self.data_row_height = self.Xattr.get_row_height(self.data_start_row)  # ?
        # 最终所需列，不需的最终去除；表头行与数据起始行之间的行，在后续读取数据时不会读取
        self.needed_col = sorted(list(needed_col))
        self.header_row = sorted(list(header_row))
        if self.header_row[-1] >= self.data_start_row:
            raise KeyError
        self.row_after_header = self.header_row[-1] + 1
        self.not_needed_col = [
            col
            for col in list(range(1, max(self.needed_col) + 1))
            if col not in self.needed_col
        ]
        self.header_max_col, self.header_max_row = max(self.needed_col), max(
            self.header_row
        )
        self.template_header_content = self.Xattr.get_range_value_df(
            0, 0, self.header_max_col, self.header_max_row
        )
        if self.header_max_row + 1 < self.data_start_row:
            self.have_seprows = True
            self.template_seprows_content = self.Xattr.get_range_value_df(
                0, self.header_max_row + 1, self.header_max_col, self.data_start_row - 1
            )
        for col in self.not_needed_col:
            self.template_header_content = self.Xattr.axising_range_value_df(
                self.template_header_content, ws_column=col
            )
        # 将样表表头之后的行，单元格属性全部清空

        for to_clear_row in self.final_excel_ws.iter_rows(min_row=self.data_start_row):
            for to_clear_cell in to_clear_row:
                self.Xattr.clear_cell_attributes(to_clear_cell)
            self.Xattr.set_row_height(to_clear_row[0].row, self.data_row_height)

    def verify_excel(self, excel_wb, excel_ws):
        """先后检验表头、跳过行、数据起始行内容是否有误

        Args:
            excel_wb (_type_): _description_
        """

        excel_item_Xattr = XPRO.Excel_attribute(excel_wb, excel_ws)
        excel_item_ws = excel_ws
        excel_item_header_content = excel_item_Xattr.get_range_value_df(
            0, 0, self.header_max_col, self.header_max_row
        )

        header_verify_flag, header_verify_output = verify_df(
            self.template_header_content, excel_item_header_content
        )
        if not header_verify_flag:
            return False, error_to_info("表头", header_verify_output)
        # if self.have_seprows:
        #     excel_item_seprows_content=excel_item_Xattr.get_range_value_df(0,self.header_max_row+1,self.header_max_col,self.data_start_row-1)
        #     excel_item_seprows_content
        # 和self.template_seprows_content对比
        # 进一步：可能不全是string，还有数字之类的，要另外写函数
        else:
            return True, ""

    def update_excel_files(self, is_verified, file_name):
        """
        验证并更新Excel文件列表。

        :param is_verified: 布尔值，表示文件是否验证通过。
        :param file_name: 字符串，表示待验证的文件名。
        """
        # 首先检查文件名是否在原始文件组中
        if file_name in self.original_excel_group:
            # 如果验证通过，将文件名和对应的数据流添加到验证通过的文件列表中
            if is_verified:
                # 从original_excel_group提取数据流，并添加到verified_excel_files中
                data_stream = self.original_excel_group[file_name]
                self.verified_excel_group[file_name] = data_stream
                
            # 不论验证是否通过，都从原始文件组中移除该文件名
            del self.original_excel_group[file_name]
            # print("len(self.original_excel_group)",len(self.original_excel_group))
            # print("len(self.verified_excel_group)",len(self.verified_excel_group))
        else:
            print(f"文件 {file_name} 不在待验证列表中。")

    def biaogexinxi(self):
        1

    def merge_and_format_excels(self) -> IO[bytes]:
        """
        将所有处理后符合格式的表格批量合并，并根据样表设置格式
        输出到数据流：合并后的Excel文件路径

        Parameters from stream:None

        Returns to stream:
            merged_excel (IO[bytes]):
                content: 合并后的Excel文件数据流
        """
        # TODO：
        #
        # 合并所有数据
        file_group_data = pd.DataFrame()
        for file_item_name, (
            file_item_wb,
            file_item_ws,
        ) in self.verified_excel_group.items():
            file_item_attr = XPRO.Excel_attribute(file_item_wb, file_item_ws)
            file_item_max_row = min(file_item_attr.get_max_row_col()["max_row"])
            file_item_data_num = file_item_max_row - self.data_start_row + 1
            file_item_data = file_item_attr.get_range_value_df(
                0, self.data_start_row, self.header_max_col, file_item_max_row
            ).iloc[1:, 1:]
            file_group_data = pd.concat([file_group_data, file_item_data])
            self.group_datanum_dict[file_item_name] = file_item_data_num

        self.Xattr.append_df_to_ws_from_row(file_group_data, self.data_start_row)

        # 设置样式
        # 使用列号匹配
        for col_letter, col_data_style in self.template_data_style.items():
            col_data_min_row, col_data_max_row = self.data_start_row, min(
                self.Xattr.get_max_row_col()["max_row"]
            )
            col_data_range = (
                f"{col_letter}{col_data_min_row}:{col_letter}{col_data_max_row}"
            )
            self.Xattr.modify_CertainRange_style(
                col_data_range, col_data_style, not_modify_attr=["value"]
            )

        # 选择表头时，可不选部分列，故此处将该列去除
        if self.not_needed_col:
            for del_col in self.not_needed_col:
                self.final_excel_ws.delete_cols(del_col)
        return self.Xio.stream_excel_to_frontend(self.final_excel_wb)


if __name__ == "__main__":
    from os.path import join as J

    # file_group_path=r"tests\for_concat\for_func1_acquire_excel_groups"
    file_group_path = r"tests\for_concat\for_func3_verify_excel"
    template_file_path = r"tests\for_concat\for_func2_extract_template_info"

    test_file_upload = XPRO.Excel_IO()
    file_group_dict = {}  # {文件名：文件流}
    # 逐一获取文件群文件，px读取，然后load到数据流
    for file_of_group in [J(file_group_path, j) for j in os.listdir(file_group_path)]:
        wb, ws = test_file_upload.read_excel_file(file_of_group)
        file_group_dict[get_filepath_variables(file_of_group)["file_name"]] = (
            test_file_upload.stream_excel_to_frontend(wb)
        )
    # 创建对象
    mergeExcelExporter = MergeExcelExporter()

    # 测试第一个方法：获取文件群，保存为属性，在此处将属性print出来方便看
    mergeExcelExporter.acquire_excel_group(file_group_dict)
    print("●得到的待合并文件群的文件名与wb对象的字典如下\n===========")
    print(mergeExcelExporter.original_excel_group)

    # 测试第二个方法：获取样表文件，产生已生成表头的最终文件

    ## 情况1：样表文件中没有数据
    template_no_data_wb, _ = test_file_upload.read_excel_file(
        J(template_file_path, "样表1_无数据.xlsx")
    )
    template_no_data_stream = test_file_upload.stream_excel_to_frontend(
        template_no_data_wb
    )

    mergeExcelExporter.extract_template_info("A1:AH1", "3", template_no_data_stream)
    # print("●该样表所选数据行的各单元格样式如下\n===========")
    # print(mergeExcelExporter.template_data_style)
    mergeExcelExporter.final_excel_wb.save(
        J(template_file_path, "仅含表头的总表1_基于" + "样表1_无数据.xlsx")
    )  # ?
    # 结果文件已保存至同一文件夹

    # ## 情况2：样表文件中有数据
    # template_with_data_wb,_=test_file_upload.read_excel_file(J(template_file_path,"样表2_有数据.xlsx"))
    # template_with_data_stream=test_file_upload.stream_excel_to_frontend(template_with_data_wb)

    # mergeExcelExporter.extract_template_info("A1:AH1","3",template_with_data_stream)
    # #print("●该样表所选数据行的各单元格样式如下\n===========")
    # #print(mergeExcelExporter.template_data_style)
    # mergeExcelExporter.final_excel_wb.save(J(template_file_path,"仅含表头的总表2_基于"+"样表2_有数据.xlsx"))#?
    # 结果文件已保存至同一文件夹

    verify_flags = []
    # 测试第三个方法
    for file_name, (wb, ws) in mergeExcelExporter.original_excel_group.items():
        verify_flag, verify_info = mergeExcelExporter.verify_excel(wb, ws)
        verify_flags.append(verify_flag)
        if not verify_flag:
            print(file_name, "错误", verify_info)
        else:
            print(file_name, True)
        # 第四个方法
        # mergeExcelExporter.update_excel_files()
    if set(verify_flags) == True:
        # 测试第五个方法
        merged_excel_stream = mergeExcelExporter.merge_and_format_excels()
        test_file_upload.load_workbook_from_stream(merged_excel_stream)[0].save(
            J(r"tests\for_concat\for_func5_merge_and_format_excels", "总表.xlsx")
        )
