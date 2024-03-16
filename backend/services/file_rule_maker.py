import os,sys,io
import openpyxl as px
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter,coordinate_to_tuple


"""用于导入项目中不在同一文件夹的库"""
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import utils.excel_processor as XPRO
import utils.string_processor as StringPRO
class FileRuleMaker:#进一步：考虑将Xio对象作为FileRuleMaker的属性，贯穿始终
    def __init__(self):
        #创建空对象占位，后续修改值，供整个类使用
        self.file_name=""                       #get_file_stream中修改
        self.file_stream=None                   #get_file_stream中修改
        self.excel_wb,self.excel_ws=None,None   #generate_user_rule_dict中修改
        self.Sheet_dropdowns=None               #generate_user_rule_dict中修改
        self.Xattr=None                         #generate_user_rule_dict中修改
        self.file_rule_dict=dict()              #create_final_rules_and_examples中修改
        self.final_excel_by_mode=dict()         #create_final_rules_and_examples中修改
        #一开始即创建，然后在整个类均可调用
        self.Xio=XPRO.Excel_IO()                                                    #自动创建，读写全部用这个对象读取。
        self.predefined_rules_path="rules/predefined_rules.json"                    #预定义规则文件的位置
        self.predefined_rules=XPRO.read_from_json_file(self.predefined_rules_path)  #预定义规则对象
        self.rule_choice_sepaprator=","                                             #进一步：考虑后端存储列表类型相关内容，传给前端的是join为字符串的内容，默认以英文逗号间隔同一规则内的各个选项，可用户自定义修改
        
        
    def get_file_stream        (self, 
                                excel_got: io.BytesIO, 
                                file_name: str, 
                                ) -> io.BytesIO:
        """
            从数据流接收  ：含字段的空文件+其文件名
            输出到数据流  ：转化后(若需)的含字段的文件数据流
            
            Parameters from stream:
                excel_got (excel_file): 含字段的内容空白的Excel文件
                file_name (str): 
                    content:该Excel文件的文件名
                    format :"test.xls(x)"
            Returns:
                xlsx_excel (io.BytesIO):后端读取后又输出(若需转化，则转化为xlsx)的xlsx文件数据流
        """
        self.file_name=file_name
        self.file_stream=excel_got
        if self.file_name.endswith("xlsx"):
            return self.file_stream
        elif self.file_name.endswith("xls"):
            self.file_stream=self.Xio.convert_excel_format(self.file_stream,"xls","xlsx",True)
            return self.file_stream#进一步：后端转化格式。修改self.excel_got变量
        else:raise TypeError#进一步：报错内容文本商讨。后端传输代号，前端呈现错误信息。
        
    def generate_user_rule_dict(self,
                                fields_index_col: dict) -> dict:
        """
            从数据流接收  ：用户选择的位置与字段值的对应字典
            输出到数据流  ：包含预定义规则和下拉列表的字典

            Parameters from stream:
                fields_index_col (dict):
                    content:所有字段的位置与字段值对应的字典
                    format :{"A1":"序号","A2":"姓名"}#进一步：考虑字段位置为合并单元格的情况
            Returns to stream:
                field_rules_for_choice (dict):
                    content:用户可选规则字典，包含预定义规则和下拉列表的字典
                    format :{"字段名1":{"对应列下拉列表规则":[["下拉规则1选项1","下拉规则1选项2"],
                                                            ["下拉规则2选项1","下拉规则2选项2"]],
                                        "程序预定义规则":   [["程序预定义规则1备注语","程序预定义规则1选项1","程序预定义规则1选项2"],
                                                            ["程序预定义规则2备注语","程序预定义规则2选项1","程序预定义规则2选项2"]]}
                            "字段名2":...同上}
        """
        # TODO: 
        # 将列号与单元格、字段名匹配  
        # 设定用户可选规则字典 ##转成列表，根据key排序，得到从左到右的字段的字典
        # 匹配字段名与预定义规则(调用函数) 
        # 匹配下拉列表信息(调用函数) 
        # 返回用户可选规则字典
        
        # 获取文件并转化
        if not self.file_stream:raise TypeError
        excel_got=self.file_stream
        self.file_stream.seek(0)
        
        self.excel_wb,self.excel_ws=self.Xio.load_workbook_from_stream(excel_got)
        # 读取对象并获取属性
        self.Xattr=XPRO.Excel_attribute(self.excel_wb,self.excel_ws)
        ""
        #进一步：改进字段名与预设规则的匹配方法
        # 匹配字段位置与字段单元格对象、字段值，已根据字段位置的先后sorted排序
        fields_index_col_to_cell_name={index_col:[self.excel_ws[index_col],fields_index_col[index_col]] for index_col in  sorted(fields_index_col.keys())}
        #return fields_index_col_to_cell_name
        # 设定用户可选规则字典 注：Python 3.6之后，字典是有序的
        self.Sheet_dropdowns=self.Xattr.get_dropdowns()
        Field_rules={col_index: 
            [name,dict(zip(["对应列下拉列表规则","程序预定义规则"],
                     [self.Sheet_dropdowns[col_index[0]] if col_index[0] in self.Sheet_dropdowns else [],
                      self.predefined_rules[StringPRO.best_match(name,list(self.predefined_rules.keys()))]
                      ]))] for col_index,(cell,name) in fields_index_col_to_cell_name.items() } 
            
        return Field_rules


    def create_final_rules_and_examples(self, 
                                selected_field_rules:dict) -> (dict,"io.BytesIO in dict"):
        """
            从数据流接收  ：字段名与规则对应的字典
            输出到数据流  ：字段名与最终规则和样例对应的字典，含有最终规则和样例行、最终规则下拉列表的Excel文件

            Parameters from stream:
                selected_field_rules (dict_saved_in_json_stream): 
                    content:用户确定后的规则字典
                    format :{"字段位置1":["字段名1",["最终规则选项1","最终规则选项2"],
                            "字段位置2":["字段名2",["最终规则选项1","最终规则选项3"],
                            "字段位置3":同上...}
                    
            Returns to stream:
                self.file_rule_dict (dict):
                    content:各个模式下，数据起始位置与字段名、最终规则和样例对应的字典
                    format :{"0-0":{"数据起始位置1":["字段名1",["最终规则正则表达式","最终规则样例"]]
                                    "数据起始位置2":...同上
                                    },
                             "1-1":{"数据起始位置1":["字段名1",["最终规则正则表达式","最终规则样例"]]
                                    "数据起始位置2":...同上
                                    },
                             "1-2":{"数据起始位置1":["字段名1",["最终规则正则表达式","最终规则样例"]]
                                    "数据起始位置2":...同上
                                    },
                             "2-2":{"数据起始位置1":["字段名1",["最终规则正则表达式","最终规则样例"]]
                                    "数据起始位置2":...同上
                                    },
                            }
                self.final_excel_by_mode (excel_file):
                    content:【最终规则样例行】与【最终规则下拉列表】自选含有与否的不同模式的Excel文件数据流存储在字典
                    format :{"0-0":io.BytesIO,                                    #(表示不对文件内容做修改)
                            "1-1":io.BytesIO,                                     #(表示在文件的字段下一行添加规则&样例行)
                            "1-2":io.BytesIO,                                     #(表示在文件除了表头的位置，均根据规则添加下拉列表)
                            "2-2":io.BytesIO}                                     #(表示同时添加规则&样例行和下拉列表)
        """
        
        # 一步一步地增加内容到wb，每一步几乎就是一种mode，直接存到字典
        def stream_mode_excel_to_dict(mode):
            self.final_excel_by_mode[mode]=self.Xio.stream_excel_to_frontend(self.excel_wb)
        # 未做修改时，为0-0
        stream_mode_excel_to_dict("0-0")
        selected_field_rules = {k: v for k, v in selected_field_rules.items() if v[1]}  # 去掉规则列表没有内容的字段
        final_rules_and_examples={}
       #设置规则样例行和最终规则样例dict，为1-1
        for one_index_col, (field_name, rule_list) in selected_field_rules.items():
            final_rules_and_examples[one_index_col]=[field_name,
                                                     StringPRO.generate_strict_regex_and_example(rule_list)]
            example=final_rules_and_examples[one_index_col][-1][-1]
            self.Xattr.set_validation_rules_and_example(one_index_col,field_name,rule_list,example)
        field_row=coordinate_to_tuple(one_index_col)[1]
        stream_mode_excel_to_dict("1-1")
        
        #设置下拉列表，为2-2
        self.Xattr.set_dropdowns(selected_field_rules)
        stream_mode_excel_to_dict("2-2")
        
        #删除规则行，上移一行下拉列表区域，为1-2(删除后，上移到删除行位置的单元格自动初始化值、字体、样式、下拉列表)
        rules_and_example_row=int(one_index_col[1])+1
        self.excel_ws.delete_rows(rules_and_example_row)
        #清空下拉列表，然后上移下拉列表
        self.excel_ws.data_validations.dataValidation.clear()
        self.Xattr.set_dropdowns(selected_field_rules, sep_row=1)
        stream_mode_excel_to_dict("1-2")
        
        # 将各个模式的规则dict写入self.file_rule_dict
        self.file_rule_dict={mode:
                                    {"".join([(str(int(i)+data_sep_row) if type(i)==int else i)
                                        for i in StringPRO.coordinate_from_string(rule_dict_key)
                                         ])
                                        :rule_dict_value
                                            for rule_dict_key,rule_dict_value in (final_rules_and_examples.items()) 
                                    }
                            for mode,data_sep_row in zip(["0-0","1-1","1-2","2-2"],[1,2,1,2])
                            }#每一种模式，一种规则文件，包含了数据起始行
        
        return self.file_rule_dict,self.final_excel_by_mode
    
    def save_final_files(self, 
                         saving_mode:str,
                         files_saving_path:str) -> str:#进一步，建议前端在这一步，为用户提供打开文件位置的快捷键
        """
            从数据流接收  ：excel文件保存模式，excel文件和规则文件保存路径
            本地操作      ：保存excel文件到指定目录，规则文件也自动保存在此目录#进一步：考虑 excel文件和规则文件 打包到一起的zip 到指定目录
            输出到数据流  ：文件保存成功提示
            Parameters from stream:
                saving_mode (str): 
                    content:excel文件保存模式,值为数字+“-”+数字
                    format :"0-0";(表示不对文件内容做修改)
                            "1-1";(表示在文件的字段下一行添加规则&样例行)
                            "1-2";(表示在文件除了表头的位置，均根据规则添加下拉列表)
                            "2-2";(表示同时添加规则&样例行和下拉列表)

            Returns to stream:
                saving_flag (str):
                    content:是否完成保存
                    format :"1"/"0"
        """
        try:
            excel_stream=self.final_excel_by_mode[saving_mode]
            excel_wb=self.Xio.load_workbook_from_stream(excel_stream)[0]
            XPRO.save_py_objection_to_json(self.file_rule_dict[saving_mode],os.path.join(os.path.dirname(files_saving_path),f"file_rule_of{saving_mode}.json"))
            for i,j in(self.file_rule_dict.items()):
                print(i,j,sep="\n")
            #input()
            self.Xio.save_excel(excel_wb,excel_path=files_saving_path)
            saving_flag="1"
        except:saving_flag="0"
        return saving_flag
    
if "__main__" == __name__:
    
    print("测试对xls文件的第一、二、三、四个方法，并将产生的四种模式的Excel文件和规则文件，保存到tests/for_fuker.allprocess_xls/saving_all_modes_test文件夹")
    #制作文件规则类的实例   
    Fuker=FileRuleMaker()

    #初始xls文件的文件名、目录、文件数据流等
    excel_got_path="tests/for_fuker.allprocess_xls/test_set_dropdown_and_ruleexamplerow.xls"########????
    excel_got=io.BytesIO()

    with open(excel_got_path, 'rb') as file:
        excel_got.write(file.read())
    # 重置流的位置到开始处，这样就可以从头读取
    excel_got.seek(0)
    
    #为保存最终xlsx文件至for_fuker.allprocess_xls文件夹做准备
    excel_got_variables=StringPRO.get_filepath_variables(excel_got_path)
    file_name=excel_got_variables["file_name"]
    file_basename,file_extension=excel_got_variables["file_basename"],excel_got_variables["file_extension"]
    
    # 用户确认后的规则字段位置字典，用于create_final_rules_and_examples
    selected_field_rules={'a5': ('序号', []), 'b5': ('作品题目', []), 'c5': ('参赛类别', []), 'd5': ('作品学科分类', ['理工农医类', '社会调查报告和人文社科类', '发明创造科技制作类']), 'e5': ('学科门类', ['哲学', '经济学', '法学', '教育学', '文学', '历史学', '理学', '工学', '农学', '医学', '军事学', '管理学', '艺术学']), 'f5': ('一级学科', ['哲学', '理论经济学', '应用经济学', '法学', '政治学', '社会学', '民族学', '马克思主义理论', '公安学', '教育学', '心理学', '体育学', '中国语言文学', '外国语言文学', '新闻传播学', '考古学', '中国史', '世界史', '数学', '物理学', '化学', '天文学', '地理学', '大气科学', '海洋科学', '地球物理学', '地质学', '生物学', '系统科学', '科学技术史', '生态学', '统计学', '力学', '机械工程', '光学工程', '仪器科学与技术', '材料科学与工程', '冶金工程', '动力工程及工程热物理', '电气工程', '电子科学与技术', '信息与通信工程', '控制科学与工程', '计算机科学与技术', '建筑学', '土木工程', '水利工程', '测绘科学与技术', '化学工程与技术', '地质资源与地质工程', '矿业工程', '石油与天然气工程', '纺织科学与工程', '轻工技术与工程', '交通运输工程', '船舶与海洋工程', '航空宇航科学与技术', '兵器科学与技术', '核科学与技术', '农业工程', '林业工程', '环境科学与工程', '生物医学工程', '食品科学与工程', '城乡规划学', '风景园林学', '软件工程', '生物工程', '安全科学与工程', '公安技术', '网络空间安全', '作物学', '园艺学', '农业资源与环境', '植物保护', '畜牧学', '兽医学', '林学', '水产', '草学', '基础医学', '临床医学', '口腔医学', '公共卫生与预防医学', '中医学', '中西医结合', '药学', '中药学', '特种医学', '医学技术', '护理学', '军事思想及军事历史', '战略学', '战役学', '战术学', '军队指挥学', '军事管理学', '军队政治工作学', '军事后勤学', '军事装备学', '军事训练学', '管理科学与工程', '工商管理', '农林经济管理', '公共管理', '图书情报与档案管理', '艺术学理论', '音乐与舞蹈学', '戏剧与影视学', '美术学', '设计学']), 'g5': ('作者', []), 'h5': ('是否为团队负责人', ['是', '否']), 'i5': ('性别', ['男', '女']), 'j5': ('生源地', ['北京', '天津', '上海', '重庆', '河北', '山西', '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽', '福建', '江西', '山东', '河南', '湖北', '湖南', '广东', '海南', '四川', '贵州', '云南', '陕西', '甘肃', '青海', '台湾', '内蒙古', '广西', '西藏', '宁夏', '新疆', '香港', '澳门', '其他']), 'k5': ('学号', []), 'l5': ('所在院系', ['数学科学学院', '物理学院', '化学与分子工程学院', '生命科学学院', '地球与空间科学学院', '城市与环境学院', '心理与认知科学学院', '建筑与景观设计学院', '信息科学技术学院', '工学院', '王选计算机研究所', '软件与微电子学院', '环境科学与工程学院', '软件工程国家工程研究中心', '中国语言文学系', '历史学系', '考古文博学院', '哲学系', '外国语学院', '艺术学院', '对外汉语教育学院', '歌剧研究院', '国际关系学院', '法学院', '信息管理系', '社会学系', '政府管理学院', '马克思主义学院', '教育学院', '新闻与传播学院', '体育教研部', '新媒体研究院', '教育财政科学研究所', '经济学院', '光华管理学院', '人口研究所', '国家发展研究院', '基础医学院', '药学院', '公共卫生学院', '护理学院', '医学人文学院', '医学继续教育学院', '第一医院', '人民医院', '第三医院', '口腔医院', '北京肿瘤医院', '第六医院', '深圳医院', '首钢医院', '国际医院', '滨海医院', '元培学院', '燕京学堂', '先进技术研究院', '前沿交叉学科研究院', '中国社会科学调查中心', '分子医学研究所', '科维理天文研究所', '核科学与技术研究院', '北京国际数学研究中心', '海洋研究院', '现代农学院', '人文社会科学研究院', '信息工程学院', '化学生物学与生物技术学院', '环境与能源学院', '城市规划与设计学院', '新材料学院', '汇丰商学院', '国际法学院', '人文社会科学学院']), 'm5': ('年级（如2020级本科生/硕士生/博士生）', []), 'n5': ('手机', []), 'o5': ('微信号', []), 'p5': ('邮箱', []), 'q5': ('指导教师姓名', []), 'r5': ('指导教师性别', ['男', '女']), 's5': ('指导教师所在院系', ['数学科学学院', '物理学院', '化学与分子工程学院', '生命科学学院', '地球与空间科学学院', '城市与环境学院', '心理与认知科学学院', '建筑与景观设计学院', '信息科学技术学院', '工学院', '王选计算机研究所', '软件与微电子学院', '环境科学与工程学院', '软件工程国家工程研究中心', '中国语言文学系', '历史学系', '考古文博学院', '哲学系', '外国语学院', '艺术学院', '对外汉语教育学院', '歌剧研究院', '国际关系学院', '法学院', '信息管理系', '社会学系', '政府管理学院', '马克思主义学院', '教育学院', '新闻与传播学院', '体育教研部', '新媒体研究院', '教育财政科学研究所', '经济学院', '光华管理学院', '人口研究所', '国家发展研究院', '基础医学院', '药学院', '公共卫生学院', '护理学院', '医学人文学院', '医学继续教育学院', '第一医院', '人民医院', '第三医院', '口腔医院', '北京肿瘤医院', '第六医院', '深圳医院', '首钢医院', '国际医院', '滨海医院', '元培学院', '燕京学堂', '先进技术研究院', '前沿交叉学科研究院', '中国社会科学调查中心', '分子医学研究所', '科维理天文研究所', '核科学与技术研究院', '北京国际数学研究中心', '海洋研究院', '现代农学院', '人文社会科学研究院', '信息工程学院', '化学生物学与生物技术学院', '环境与能源学院', '城市规划与设计学院', '新材料学院', '汇丰商学院', '国际法学院', '人文社会科学学院']), 't5': ('指导教师职称/职务', []), 'u5': ('指导教师电话', []), 'v5': ('指导教师电子邮箱', [])}
    
    #字段位置与名称对应的字典，用于generate_user_rule_dict
    fields_index_col={i:j[0] for i,j in selected_field_rules.items()}
    
    #第一个方法
    Fuker.get_file_stream(excel_got,file_name)
    
    #第二个方法
    Fuker.generate_user_rule_dict(fields_index_col)
    
    #第三个方法，得到的字典、得到的文件都存在属性并作为方法返回值
    output_rule_dict,output_excel_dict=Fuker.create_final_rules_and_examples(selected_field_rules)
    
    #第四个方法，此处将四种模式的Excel文件都保存了，文件名中含有其模式;规则文件前后覆盖地保存了四次，故最终只有一个规则json
    for i in """0-0
1-1
1-2
2-2""".split("\n"):
        new_file_name="allprocess_xls_"+i+"_"+file_basename+".xlsx"
        new_file_save_path=os.path.join(excel_got_variables["folder_path"],"saving_all_modes_test",new_file_name)
        print(Fuker.save_final_files(i,new_file_save_path))
    