import sys, io, json, re, os, warnings, shutil, copy
import openpyxl as px
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import range_boundaries
import win32com.client as win32
from typing import IO, List, Dict, Union

warnings.filterwarnings("ignore", category=UserWarning)

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import utils.excel_processor as XPRO
from utils.string_processor import *

# from services.file_rule_maker import FileRuleMaker


class FileValidator:
    def __init__(self):
        # 创建空对象占位，后续修改值，供整个类使用
        self.file_name = ""  # get_files_stream中修改
        self.file_stream = None  # get_files_stream中修改
        self.excel_wb, self.excel_ws = None, None  # validate_filled_excel中修改
        self.Xattr = None  # validate_filled_excel中修改
        self.range_and_rule = (
            dict()
        )  # 在get_files_stream中获得规则值，在validate_filled_excel中获得区域

        # 一开始即创建，然后在整个类均可调用
        self.Xio = XPRO.Excel_IO()  # 自动创建，读写全部用这个对象读取。
        self.error_cell_style = {
            "fill": PatternFill(
                start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
            ),
            "font": Font(strike=True),
        }
        self.original_cell_style = {}  # 从文件读取正确单元格样式

    def get_files_stream(
        self,
        excel_got: io.BytesIO,
        file_name: str,
        final_rules_and_examples: dict,
    ) -> io.BytesIO:
        """
        从数据流接收  ：用户填写完毕的文件+其文件名+从json读取过来的规则字典
        输出到数据流  ：转化后(若需)的文件数据流

        Parameters from stream:
            excel_got (excel_file): 用户填写完毕的Excel文件
            file_name (str):
                content:该Excel文件的文件名
                format :"test.xls(x)
            file_rule_dict (dict):
                content:字段名与最终规则和样例对应的字典
                format :{"数据起始位置1":["字段名1",["最终规则正则表达式","最终规则样例"]]
                            "数据起始位置2":...同上}
        Returns:
            xlsx_excel (io.BytesIO):后端读取后又输出(若需转化，则转化为xlsx)的xlsx文件数据流

        """
        # 创建空对象占位，后续修改值，供整个类使用
        self.file_name = ""  # get_files_stream中修改
        self.file_stream = None  # get_files_stream中修改
        self.excel_wb, self.excel_ws = None, None  # validate_filled_excel中修改
        self.Xattr = None  # validate_filled_excel中修改
        self.range_and_rule = (
            dict()
        )  # 在get_files_stream中获得规则值，在validate_filled_excel中获得区域

        # 一开始即创建，然后在整个类均可调用
        self.Xio = XPRO.Excel_IO()  # 自动创建，读写全部用这个对象读取。
        self.error_cell_style = {
            "fill": PatternFill(
                start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
            ),
            "font": Font(strike=True),
        }
        self.original_cell_style = {}  # 从文件读取正确单元格样式
        # 读取excel文件和文件名
        self.file_name = file_name
        self.file_stream = excel_got

        # 读取数据起始位置、字段名、规则正则表达式、规则样例到dict，该字段数据区域在下一步validate_filled_excel中读出
        self.range_and_rule = final_rules_and_examples

        if self.file_name.endswith("xls"):
            self.file_stream = self.Xio.convert_excel_format(
                self.file_stream, "xls", "xlsx", True
            )
        self.excel_wb, self.excel_ws = self.Xio.load_workbook_from_stream(
            self.file_stream
        )
        self.file_stream.seek(0)
        self.Xattr = XPRO.Excel_attribute(self.excel_wb, self.excel_ws)
        # self.original_cell_style.update(self.Xattr.get_row_attributes())
        return self.file_stream

    def validate_filled_excel(
        self, filled_excel_file: io.BytesIO
    ) -> Union[dict, IO[bytes]]:
        """
        验证已填写的Excel文件是否符合规则，并标记不符合的单元格。

        Parameters:
            filled_excel_file (file): 已填写的Excel文件。

        Returns:
            wrong_cells_info (dict)  :
                        content:key:错误单元格位置；value:对应规则转化成的选项提示语
                        format :{"错误单元格1位置": "该单元格选项可以为：是，否",
                                 "错误单元格2位置": "..."}

            validated_excel (io.BytesIO):错误单元格样式修改后的excel数据流
        """

        if not filled_excel_file:
            raise TypeError

        # 读取对象并获取属性
        filled_excel_file.seek(0)
        self.excel_wb, self.excel_ws = self.Xio.load_workbook_from_stream(
            filled_excel_file
        )
        self.Xattr = XPRO.Excel_attribute(self.excel_wb, self.excel_ws)

        # 在load 的json文件dict每个key的value基础上，再加上了该规则适用的数据范围
        data_end_row = min(self.Xattr.get_max_row_col()["max_row"])
        range_and_rule = copy.deepcopy(self.range_and_rule)
        for data_begin_cell in range_and_rule:
            data_col, data_begin_row = coordinate_from_string(data_begin_cell)
            data_range = f"{data_col}{data_begin_row}:{data_col}{data_end_row}"
            range_and_rule[data_begin_cell].append(data_range)

        # print(range_and_rule)
        def validate_all_data(range_and_rule=range_and_rule):
            # 获得错误单元格的位置key 规则、样例
            error_cells_info = {}
            for data_field, rule_and_example, data_range in range_and_rule.values():
                data_rule, data_example = rule_and_example
                # print(self.excel_ws[data_range])
                for cell in self.excel_ws[data_range]:
                    # cell对象使用区域遍历时，形如(<Cell 'Sheet1'.A1>,)
                    cell = cell[0]
                    if not match_with_regex(data_rule, cell.value):
                        error_cells_info[cell.coordinate] = [data_rule, data_example]
            return error_cells_info

        # 首先检验获得当前所有错误单元格
        current_error_cells_info = validate_all_data()
        previous_error_cells = copy.copy(list(self.original_cell_style.keys()))
        # for i,style__ in self.original_cell_style.items():
        #     print("saved_cell",i)
        #     print("saved_style__",style__['fill'])
        # 然后检验之前的错误单元格是否错误，若正确则字典去除该单元格并恢复样式，若错误不用管
        for previous_error_cell in previous_error_cells:
            if previous_error_cell not in current_error_cells_info:
                validated_cell_style = copy.deepcopy(
                    self.original_cell_style[previous_error_cell]
                )
                del self.original_cell_style[previous_error_cell]
                self.Xattr.modify_cell_style(previous_error_cell, validated_cell_style)
                # print("●previous_error_cell","对了",previous_error_cell)
                # print("填充改为",self.Xattr.get_cell_attributes(previous_error_cell)[previous_error_cell]['fill'])
                # print(previous_error_cell in current_error_cells_info)

        # 最后检验当前错误单元格是否上一次错了，错了则不改样式，否则修改样式为错误样式并保存之前样式；两种情况均产生错误单元格提示文本到字典
        wrong_cells_info = {}
        for current_error_cell, (
            data_rule,
            data_example,
        ) in current_error_cells_info.items():
            if current_error_cell not in self.original_cell_style:
                self.original_cell_style.update(
                    self.Xattr.get_cell_attributes(current_error_cell, ["value"])
                )

                self.Xattr.modify_cell_style(current_error_cell, self.error_cell_style)
            wrong_cells_info[current_error_cell] = (
                f"{transform_pattern_to_description(data_rule)}\n例如可填写：{data_example}"
            )

        validated_excel_stream = self.Xio.stream_excel_to_frontend(self.excel_wb)
        return wrong_cells_info, validated_excel_stream


if __name__ == "__main__":

    # 创造对象
    Favor = FileValidator()

    # 读取excel文件、规则json文件，是第一个方法的参数

    ##初始xlsx文件的文件名、目录、文件数据流等
    excel_got_path = r"tests\for_favor_func1_2_3\登记表+税远志.xls"

    ##用于后续输出文件的参数，输出文件用于方便直观感受程序结果
    excel_got_variables = get_filepath_variables(excel_got_path)
    file_name = excel_got_variables["file_name"]
    file_basename, file_extension = (
        excel_got_variables["file_basename"],
        excel_got_variables["file_extension"],
    )

    ##将文件读取后写入数据流，作为第一个方法的参数
    excel_got = io.BytesIO()
    with open(excel_got_path, "rb") as file:
        excel_got.write(file.read())

    ## 重置流的位置到开始处，这样就可以从头读取
    excel_got.seek(0)
    file_rule = XPRO.read_from_json_file(
        os.path.join(excel_got_variables["folder_path"], "file_rule_of0-0.json")
    )

    # 测试第一个方法
    func1_file_name = "after_func1_" + file_basename + ".xlsx"
    func1_file_save_path = os.path.join(
        excel_got_variables["folder_path"], func1_file_name
    )

    excel_stream = Favor.get_files_stream(excel_got, file_name, file_rule)

    print(f"func1:\n程序已转化前端传来的excel")
    ## 保存return值，方便感受结果，即方法一输出的excel文件
    excel_wb_func1 = Favor.Xio.load_workbook_from_stream(excel_stream)[0]
    Favor.Xio.save_excel(excel_wb_func1, excel_path=func1_file_save_path)

    # 测试第二个方法(样例文件中D6 D7 D8都写错了，程序将予以标红)
    for i in range(1, 7):
        print(f"●func2测试第{i}次")
        func2_file_name = f"after_func2_change_d{5+i}" + file_basename + ".xlsx"
        func2_file_save_path = os.path.join(
            excel_got_variables["folder_path"], func2_file_name
        )
        Favor.Xattr.excel_ws[f"D{5+i}"].value = "发明创造科技制作类"
        error_info, excel_ = Favor.validate_filled_excel(
            Favor.Xio.stream_excel_to_frontend(Favor.Xattr.excel_wb)
        )
        error_list = list(error_info.keys())
        print(f"func2:\n程序发现表格内有错误单元格如下:\n{error_list}")
        ## 保存return值，方便感受结果，即方法二输出的excel文件
        excel_wb_func2 = Favor.Xio.load_workbook_from_stream(excel_)[0]
        Favor.Xio.save_excel(excel_wb_func2, excel_path=func2_file_save_path)
        print(f"saving to {func2_file_save_path}")

    # # 全部修改正确
    # for j in range(1,3):
    #     func3_file_name=f"after_func2_change_d8_第{j}次"+file_basename+".xlsx"
    #     func3_file_save_path=os.path.join(excel_got_variables["folder_path"],func3_file_name)
    #     # 修改错误值后使用第二个方法验证有无错误值
    #     excel_wb_to_correct=excel_wb_func2
    #     excel_ws_to_correct=excel_wb_to_correct.worksheets[0]
    #     for cell in [excel_ws_to_correct[index_col] for index_col in error_info]:
    #         cell.value="发明创造科技制作类"
    #     excel_after_correct=Favor.Xio.stream_excel_to_frontend(excel_wb_to_correct)

    # # 后端检验发现无错误值，然后假设前端用户确认无误，并将输出路径设置为func3_file_save_path，后端接收并保存excel
    # error_info,validated_excel_stream=Favor.validate_filled_excel(excel_after_correct)
    # if not error_info:
    #     print(f"func3:\n经程序检验，表格无误")
    #     print("收到前端用户发送的保存地址，最终excel已保存")
    #     ## 保存最后的excel
    #     Favor.Xio.load_workbook_from_stream(validated_excel_stream)[0].save(func3_file_save_path)
