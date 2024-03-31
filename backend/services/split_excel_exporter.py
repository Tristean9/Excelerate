import os,sys,io 
import pandas as pd
import openpyxl as px
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter,coordinate_to_tuple,range_boundaries,column_index_from_string
from typing import IO, List, Dict, Union

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import utils.excel_processor as XPRO
from   utils.string_processor import *

class SplitExcelExporter:
    # 此方法无需预先手动按照参考列排序，pd一键实现
    def __init__(self):
        """
        初始化方法
        创建用于处理拆分操作的内部变量
        """
        # TODO: 初始化内部变量
        self.excel_wb,self.excel_ws=None,None#load_excel_parameters
        self.data_start_row,self.reference_column=0,""#load_excel_parameters
        self.split_files={}#
        self.Xio=XPRO.Excel_IO()
        self.Xattr=None#
        self.data_start_row=0#load_excel_parameters
        self.header_cells_attr={}#load_excel_parameters
        self.data_row_style={}#load_excel_parameters
        self.all_data_df=None#load_excel_parameters
        self.split_data_df={}
        self.split_data_excel={}
        
    def load_excel_parameters(self, 
                              excel_stream: IO[bytes],
                              data_start_row: Union[str, int],
                              reference_column: str) -> None:
        """
        加载前端自行转化为xlsx且用户已确认的Excel文件 #进一步：是否前端打开后发后端？需不需要不经手前端实现文件的纯净化？
        从数据流中读取Excel文件、数据开始行和拆分参考的列以进行处理

        Parameters:
            excel_stream (IO[bytes]): 包含Excel文件内容的数据流
            data_start_row (Union[str, int]): 数据开始的行号
            reference_column (str): 用于拆分的参考列的列号letter

        Returns to stream: None
        """
        # TODO: 
        # 实现前端确认后的Excel文件的加载
        # 获取数据行的格式
        
        self.excel_wb,self.excel_ws=self.Xio.load_workbook_from_stream(excel_stream)
        self.Xattr=XPRO.Excel_attribute(self.excel_wb,self.excel_ws)
        self.data_start_row=data_start_row
        self.reference_column=reference_column
        
        # 获取表头、样表样式？表头要不循环往复利用？???????进一步
        if self.data_start_row>1:
            for header_row_num in range(1,self.data_start_row+1):
                self.header_cells_attr.update(self.Xattr.get_row_attributes(header_row_num))
        self.data_row_style.update(self.Xattr.get_row_attributes(data_start_row,["hyperlink"]))#,"value"
        print("number_format",self.data_row_style["C3"]["value"])
        
        
        data_boundary_dict=self.Xattr.get_max_row_col()
        self.data_max_row,self.data_max_col=min(data_boundary_dict["max_row"]),min(data_boundary_dict["max_col"])
        self.all_data_df=self.Xattr.get_range_value_df(0,self.data_start_row,self.data_max_col,self.data_max_row).iloc[1:,1:]

        
    def split_worksheet(self) -> Dict[str,int]:
        """
        按照特定列的值拆分Excel工作表
        根据设置的数据开始行和拆分参考列执行拆分操作

        Parameters from stream: None

        Returns to stream:
            split_files_info (Dict[int]):
                content: {"依据列内容值":[行数]}
        """
        df=self.all_data_df

        # 保留原始列索引
        original_columns = df.columns
        reference_column_index=column_index_from_string(self.reference_column)
        print("●",list(original_columns),reference_column_index)
        # 从数据开始行号开始拆分DataFrame
        self.split_data_df = {value: df.loc[df[reference_column_index] == value, original_columns]
                     for value in df[reference_column_index].unique()}

        # TODO: 
        # 先用pd拆分成若干个df #进一步：时间格式是否变化检验，和合并功能也有关。
        # 再遍历写入ws，并赋样式。
        
        return {value:split_df.shape[0] for value,split_df in self.split_data_df.items()}
    
    def zip_split_files(self, 

                         file_name_dict: Dict[str, str]) -> IO[bytes]:
        #样式
        """

        保存拆分后的文件#未改
        将拆分后获取的数据流保存到excel，存储在zip数据流中

        Parameters from stream:
            file_name_dict(Dict[str,str])
                content: 拆分依据字段的具体值:对应字段值拆分后的文件名（可加前缀、后缀等）
                {"信息管理系":"1-信管"}

        Returns to stream:
            file_paths (List[str]):
                content: 拆分后的文件路径列表
        """

        for file_name,data_df in self.split_data_df.items():
            clear_row_nums=max(self.Xattr.get_max_row_col()["max_row"])-self.data_start_row+1
            self.excel_ws.delete_rows(self.data_start_row,clear_row_nums)
            self.Xattr.append_df_to_ws_from_row(data_df,self.data_start_row)
            
            # 修改后，匹配样式
            ## 先找区域
            data_col_ranges={}
            data_boundary_dict=self.Xattr.get_max_row_col()
            data_max_col=min(data_boundary_dict["max_col"])
            data_max_row=min(data_boundary_dict["max_row"])
            for col  in range(1,data_max_col+1):
                col_letter=get_column_letter(col)
                data_col_range=f"{col_letter}{self.data_start_row}:{col_letter}{data_max_row}"
                data_col_ranges[f"{col_letter}{self.data_start_row}"]=data_col_range
            
            # 再设置样式 
            for col_fistr_cell,col_range in data_col_ranges.items():
                self.Xattr.modify_CertainRange_style(col_range,self.data_row_style[col_fistr_cell])
                
                self.split_data_excel[file_name_dict[file_name]]=self.Xio.stream_excel_to_frontend(self.excel_wb)
        
        return XPRO.stream_files_to_zip(self.split_data_excel)


if __name__ == "__main__":
    from os.path import join as J
    
    original_file_path=r"tests\for_concat\for_func5_merge_and_format_excels\总表.xlsx"
    split_file_path=r"tests\for_split"
    
    original_file_wb=px.load_workbook(original_file_path)
    original_file_stream=XPRO.Excel_IO().stream_excel_to_frontend(original_file_wb)

    splitExcelExporter=SplitExcelExporter()
    splitExcelExporter.load_excel_parameters(original_file_stream,3,"E")
    print(splitExcelExporter.split_worksheet())
    # for split_reference,split_df in (splitExcelExporter.split_data_df).items():
    #     split_df.to_excel(J(split_file_path,split_reference+".xlsx"))
    # 压缩
    zip_buffer = splitExcelExporter.zip_split_files({'前沿交叉学科研究院': "1-叉院.xlsx", '地球与空间科学学院': "2-地空.xlsx", '城市与环境学院': "3-城环.xlsx"})
    
    zip_buffer.seek(0)
    file_path=os.path.join(split_file_path,"拆分后.zip")
    # 这里仅仅是为了演示，实际上应该是将zip_buffer传送给前端进行下载
    with open(file_path, "wb") as f:
        f.write(zip_buffer.getvalue())
    