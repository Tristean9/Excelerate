import difflib,re,random,os
import pandas as pd
def get_filepath_variables(excel_got):
    """返回excel_got的folder_path,file_basename,file_extension,file_name的键值对字典"""
    folder_path=os.path.dirname(excel_got)
    file_basename,file_extension=os.path.splitext(os.path.basename(excel_got))
    file_name=os.path.basename(excel_got)
    return {name_:variable_ for name_,variable_ in zip("folder_path file_basename file_extension file_name".split(),[folder_path,file_basename,file_extension,file_name])}
def have_common_characters(str1, str2):
    return bool(set(str1) & set(str2))
def best_match(target, options):
    """在选项中找到与目标字符串最接近的字符串。
    :param target: 目标字符串
    :param options: 字符串列表，用于与目标进行匹配
    :return: 匹配度最高的字符串
    """
    # 获取匹配度最高的字符串,
    matches = difflib.get_close_matches(target, options, n=1, cutoff=0.0)
    # 若没有或者甚至无共同字符，返回""
    if not matches:return ""
    elif not have_common_characters(target,matches[0]):return ""
    # 如果有匹配的，返回第一个（最佳匹配），否则返回None
    
    else:return matches[0]

def generate_strict_regex_and_example(input_list):
    # 使用 '^' 和 '$' 生成严格匹配列表中任一项的正则表达式
    regex_pattern = r'^(?:' + '|'.join(re.escape(item) for item in input_list) + r')$'
    
    # 从列表中随机选择一个样例
    random_example = random.choice(input_list)
    
    return [regex_pattern, random_example]

# 使用正则表达式
def match_with_regex(regex, string_to_test):
    return re.match(regex, string_to_test) is not None
def check_strings(a, b):#范围比较宽 #进一步：修改
    # 移除字符串中的所有空白符（包括空格、换行符等）
    if a== "__SPECIAL_VALUE__" or b=="__SPECIAL_VALUE__":return True
    stripping = lambda x:("".join(x.split())).replace("（","(").replace("）",")").upper()
    stripped_a = stripping(a)
    stripped_b = stripping(b)
    return stripped_a in stripped_b or stripped_b in stripped_a
def check_cell_value(value1, value2):
    # 处理 NaN 情况
    if pd.isnull(value1) and pd.isnull(value2):
        return True
    # 确保两个值的类型相同
    try:
        if isinstance(value1, float) or isinstance(value2, float):
            # 尝试转换为float进行比较
            return float(value1) == float(value2)
        elif isinstance(value1, int) or isinstance(value2, int):
            # 尝试转换为int进行比较
            return int(value1) == int(value2)
        else:
            # 默认作为str处理
            return check_strings(value1,value2)
    except ValueError:
        # 如果转换失败，直接比较原始值
        return value1 == value2
import pandas as pd

# 假设的 check_cell_value 函数
def check_cell_value(value1, value2):
    return value1 == value2

def verify_df(df1, df2):
    if df1.shape != df2.shape:
        return False, {}
    # 去除填充的上方左方单元格
    df1=df1.iloc[1:,1:]
    df2=df2.iloc[1:,1:]
    wrong_rows = []
    wrong_cells = []
    wrong_cols = []
    col_pairs = []

    # 比较行
    for index, (row1, row2) in enumerate(zip(df1.iterrows(), df2.iterrows())):
        row_index, rowData1 = row1
        _, rowData2 = row2
        row_compare = [check_cell_value(cell1, cell2) for cell1, cell2 in zip(rowData1, rowData2)]

        if all(row_compare):
            continue
        elif not any(row_compare):
            return False, {"row_index": row_index}
        else:
            wrong_rows.append(index)
            wrong_cells.extend([(index, col_index) for col_index, correct in enumerate(row_compare) if not correct])
    if wrong_rows==[]:return True,{}
    """# 检查错误的单元格是否集中在特定的列 #进一步。不是简单交换、交换后df相等
    wrong_cols = list(set([cell[1] for cell in wrong_cells]))

    # 尝试列交换
    if len(wrong_cols) % 2 == 0:
        for i in range(0, len(wrong_cols), 2):
            col1, col2 = wrong_cols[i], wrong_cols[i+1]
            df2_copy = df2.copy()
            df2_copy.iloc[:, [col1, col2]] = df2_copy.iloc[:, [col2, col1]]
            if df2_copy.equals(df1):
                col_pairs.append((col1, col2))
        if col_pairs:
            return False, {"change_col": col_pairs}"""

    # 如果没有正确的列交换，返回错误的单元格坐标
    return False, {"cell_coord": wrong_cells}


if "codes from openpyxl.util.cell":
    # Copyright (c) 2010-2020 openpyxl

    """
    Collection of utilities used within the package and also available for client code
    """
    import re
    from string import digits
    from openpyxl.utils.exceptions import CellCoordinatesException

    # constants
    COORD_RE = re.compile(r'^[$]?([A-Za-z]{1,3})[$]?(\d+)$')
    COL_RANGE = """[A-Z]{1,3}:[A-Z]{1,3}:"""
    ROW_RANGE = r"""\d+:\d+:"""
    RANGE_EXPR = r"""
    [$]?(?P<min_col>[A-Za-z]{1,3})?
    [$]?(?P<min_row>\d+)?
    (:[$]?(?P<max_col>[A-Za-z]{1,3})?
    [$]?(?P<max_row>\d+)?)?
    """
    ABSOLUTE_RE = re.compile('^' + RANGE_EXPR +'$', re.VERBOSE)
    SHEET_TITLE = r"""
    (('(?P<quoted>([^']|'')*)')|(?P<notquoted>[^'^ ^!]*))!"""
    SHEETRANGE_RE = re.compile("""{0}(?P<cells>{1})(?=,?)""".format(
        SHEET_TITLE, RANGE_EXPR), re.VERBOSE)


    def get_column_interval(start, end):
        """
        Given the start and end columns, return all the columns in the series.

        The start and end columns can be either column letters or 1-based
        indexes.
        """
        if isinstance(start, str):
            start = column_index_from_string(start)
        if isinstance(end, str):
            end = column_index_from_string(end)
        return [get_column_letter(x) for x in range(start, end + 1)]



    def coordinate_from_string(coord_string):
        """Convert a coordinate string like 'B12' to a tuple ('B', 12)"""
        match = COORD_RE.match(coord_string)
        if not match:
            msg = f"Invalid cell coordinates ({coord_string})"
            raise CellCoordinatesException(msg)
        column, row = match.groups()
        row = int(row)
        if not row:
            msg = f"There is no row 0 ({coord_string})"
            raise CellCoordinatesException(msg)
        return column, row



    def absolute_coordinate(coord_string):
        """Convert a coordinate to an absolute coordinate string (B12 -> $B$12)"""
        m = ABSOLUTE_RE.match(coord_string)
        if not m:
            raise ValueError(f"{coord_string} is not a valid coordinate range")

        d = m.groupdict('')
        for k, v in d.items():
            if v:
                d[k] = f"${v}"

        if d['max_col'] or d['max_row']:
            fmt = "{min_col}{min_row}:{max_col}{max_row}"
        else:
            fmt = "{min_col}{min_row}"
        return fmt.format(**d)



    def _get_column_letter(col_idx):
        """Convert a column number into a column letter (3 -> 'C')

        Right shift the column col_idx by 26 to find column letters in reverse
        order.  These numbers are 1-based, and can be converted to ASCII
        ordinals by adding 64.

        """
        # these indicies corrospond to A -> ZZZ and include all allowed
        # columns
        if not 1 <= col_idx <= 18278:
            raise ValueError("Invalid column index {0}".format(col_idx))
        letters = []
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx, 26)
            # check for exact division and borrow if needed
            if remainder == 0:
                remainder = 26
                col_idx -= 1
            letters.append(chr(remainder+64))
        return ''.join(reversed(letters))


    _COL_STRING_CACHE = {}
    _STRING_COL_CACHE = {}
    for i in range(1, 18279):
        col = _get_column_letter(i)
        _STRING_COL_CACHE[i] = col
        _COL_STRING_CACHE[col] = i


    def get_column_letter(idx,):
        """Convert a column index into a column letter
        (3 -> 'C')
        """
        try:
            return _STRING_COL_CACHE[idx]
        except KeyError:
            raise ValueError("Invalid column index {0}".format(idx))



    def column_index_from_string(str_col):
        """Convert a column name into a numerical index
        ('A' -> 1)
        """
        # we use a function argument to get indexed name lookup
        try:
            return _COL_STRING_CACHE[str_col.upper()]
        except KeyError:
            raise ValueError("{0} is not a valid column name".format(str_col))



    def range_boundaries(range_string):
        """
        Convert a range string into a tuple of boundaries:
        (min_col, min_row, max_col, max_row)
        Cell coordinates will be converted into a range with the cell at both end
        """
        msg = "{0} is not a valid coordinate or range".format(range_string)
        m = ABSOLUTE_RE.match(range_string)
        if not m:
            raise ValueError(msg)

        min_col, min_row, sep, max_col, max_row = m.groups()

        if sep:
            cols = min_col, max_col
            rows = min_row, max_row

            if not (
                all(cols + rows) or
                all(cols) and not any(rows) or
                all(rows) and not any(cols)
            ):
                raise ValueError(msg)

        if min_col is not None:
            min_col = column_index_from_string(min_col)

        if min_row is not None:
            min_row = int(min_row)

        if max_col is not None:
            max_col = column_index_from_string(max_col)
        else:
            max_col = min_col

        if max_row is not None:
            max_row = int(max_row)
        else:
            max_row = min_row

        return min_col, min_row, max_col, max_row



    def rows_from_range(range_string):
        """
        Get individual addresses for every cell in a range.
        Yields one row at a time.
        """
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        rows = range(min_row, max_row + 1)
        cols = [get_column_letter(col) for col in range(min_col, max_col + 1)]
        for row in rows:
            yield tuple('{0}{1}'.format(col, row) for col in cols)



    def cols_from_range(range_string):
        """
        Get individual addresses for every cell in a range.
        Yields one row at a time.
        """
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        rows = range(min_row, max_row+1)
        cols = (get_column_letter(col) for col in range(min_col, max_col+1))
        for col in cols:
            yield tuple('{0}{1}'.format(col, row) for row in rows)



    def coordinate_to_tuple(coordinate):
        """
        Convert an Excel style coordinate to (row, colum) tuple
        """
        for idx, c in enumerate(coordinate):
            if c in digits:
                break
        col = coordinate[:idx].upper()
        row = coordinate[idx:]
        return int(row), _COL_STRING_CACHE[col]



    def range_to_tuple(range_string):
        """
        Convert a worksheet range to the sheetname and maximum and minimum
        coordinate indices
        """
        m = SHEETRANGE_RE.match(range_string)
        if m is None:
            raise ValueError("Value must be of the form sheetname!A1:E4")
        sheetname = m.group("quoted") or m.group("notquoted")
        cells = m.group("cells")
        boundaries = range_boundaries(cells)
        return sheetname, boundaries



    def quote_sheetname(sheetname):
        """
        Add quotes around sheetnames if they contain spaces.
        """
        if "'" in sheetname:
            sheetname = sheetname.replace("'", "''")

        sheetname = u"'{0}'".format(sheetname)
        return sheetname

"""# 示例使用match
options_list = ["填写", "日", "院系日名称", "院系"]
target_string = "日期"

# 输出匹配度最高的字符串
best_match_string = best_match(target_string, options_list)
print(best_match_string)
"""

if "__main__" == __name__:
    # 示例使用re
    my_list = ['apple', 'banana', 'cherry']
    regex, example = generate_strict_regex_and_example(my_list)

    # 验证正则表达式
    test_string = 'banana'
    if match_with_regex(regex, test_string):
        print(f'The string "{test_string}" is an exact match in the list.')
    else:
        print(f'The string "{test_string}" does not exactly match any item in the list.')

    print(f'Regex: {regex}')
    print(f'Random example: {example}')