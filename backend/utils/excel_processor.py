import openpyxl as px
import pandas as pd
import numpy as np
from copy import copy,deepcopy
import io,json,re,os,warnings,shutil,sys,zipfile
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries, get_column_letter
import win32com.client as win32
import pythoncom
warnings.filterwarnings("ignore", category=UserWarning)

"""用于导入项目中不在同一文件夹的库"""
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


import utils.string_processor as StringPRO

class Excel_IO:
    def __init__(self):
        # Excel格式映射
        self.FORMATS = {'xls': 56,    'xlsx': 51}
        self.temp_path="tmp/"
    def read_excel_file(self, excel_path, sheet_index=0):
        """openpyxl读取某路径的excel文件,有点害人,返回的是wb和ws的tuple,略微合理。"""
        try:
            excel_wb = px.load_workbook(excel_path, data_only=True)
            # 判断是否存在名为"Evaaluation Version"的工作表
            if "Evaluation Version" in excel_wb.sheetnames:
                sheet_index+=1
            excel_ws = excel_wb.worksheets[sheet_index]
            return (excel_wb,excel_ws)
        except IOError as e:
            print(f"An error occurred during reading: {e}")
            # Handle the exception as needed
            return None

    def load_workbook_from_stream(self,excel_stream, sheet_index=0):
        """openpyxl读取某数据流的excel文件,有点害人,返回的是wb和ws的tuple,略微合理。"""
        if 1:#try:
            # 读取流中的内容为二进制数据
            excel_data = excel_stream.read()
            # 使用BytesIO创建一个类似文件的对象
            excel_bytes = io.BytesIO(excel_data)
            return self.read_excel_file(excel_bytes)

    
    def save_excel(self, excel_wb, excel_path):
        """openpyxl传输wb对象到excel文件"""
        try:
            excel_wb.save(excel_path)
        except IOError as e:
            print(f"An error occurred during saving: {e}")

            
    def stream_excel_to_frontend(self, excel_wb):
        """openpyxl传输wb对象到excel数据流"""
        try:
            # 创建一个BytesIO对象来保存Excel文件
            excel_stream = io.BytesIO()
            # 将workbook保存到这个BytesIO流中
            excel_wb.save(excel_stream)

            # 重置流的位置到开始处，这样就可以读取它的内容了
            excel_stream.seek(0)

            # 返回流对象
            return excel_stream
        except IOError as e:
            print(f"An error occurred during streaming: {e}")
            return None

    def convert_excel_format(self,input_bytes, src_format, dst_format,save_dst=True):
        """根据参数将数据流中的excel格式进行转化，并输出为数据流,默认在tmp文件夹中产生的临时文件"""
        
        # 清理之前的临时文件
        clear_directory(self.temp_path)
        
        pythoncom.CoInitialize()
        
        # 确保源格式和目标格式是受支持的
        if src_format not in self.FORMATS or dst_format not in self.FORMATS:
            raise ValueError('Unsupported format specified.')

        src_tempfile_path=os.path.abspath(os.path.join(self.temp_path,f"temp.{src_format}"))
        dst_tempfile_path=os.path.abspath(os.path.join(self.temp_path,f"temp.{dst_format}"))
        
        # 创建 Excel 对象
        excel = win32.DispatchEx('Excel.Application')
        excel.Visible = False  # 不显示Excel界面
        
        # 创建输出流
        output_io = io.BytesIO()

        # 将输入BytesIO对象中的内容写入临时源文件
        with open(src_tempfile_path, "wb") as temp_file:
            temp_file.write(input_bytes.getvalue())
        # 打开源文件
        try:
            workbook = excel.Workbooks.Open(os.path.abspath(src_tempfile_path))

            # 另存为目标格式的文件
            workbook.SaveAs(dst_tempfile_path, FileFormat=self.FORMATS[dst_format])
        
            
        finally:
            workbook.Close(False)#True改为False了
            excel.Quit()#去掉了后面quit

        # 读取目标文件到BytesIO对象
        with open(dst_tempfile_path, "rb") as temp_file:
            output_io.write(temp_file.read())

        # 清理临时文件
        os.remove(src_tempfile_path)
        
        #在最后的保存excel步骤，可先保留文件至temp文件夹，再传输到用户选择的文件夹
        if not save_dst:
            os.remove(dst_tempfile_path)

        # # 关闭 Excel 进程
        # excel.Application.Quit()

        # 设置输出流的指针回到起始位置，以便于读取
        output_io.seek(0)
        pythoncom.CoUninitialize()
        return output_io


class Excel_attribute:
    """目前只考虑了一个工作簿&其一个工作表的修改，进一步：无法实现多个工作表同时修改"""
    def __init__(self, excel_wb=None , excel_ws=None):
        """类无传输值分别表示创建新wb、读取wb第一个工作表"""
        if excel_wb is None:
            self.excel_wb = px.Workbook()
            self.excel_ws = self.excel_wb.active
        else:
            self.excel_wb = excel_wb
            self.excel_ws = excel_ws if excel_ws is not None else excel_wb.worksheets[0]
    
    def get_some_axis_cells(self,index,value_only=True):
        """获取某一行/列的单元格，依据参数返回单元格对象或值的list,字母和数字分别指代列和行"""
        
        transform_cell=lambda cell:cell.value if value_only==True else cell
        excel_field=[transform_cell(cell) for cell in self.excel_ws[index]]
        return excel_field
    
    def get_range_value_df(self, min_col, min_row, max_col, max_row,excel_ws=None):#需要获取除了self.excel_ws之外的其它单元格吗
        """获取指定区域内的单元格值，存储到df,可以行列索引#保护最大行、最大列"""
        cells = []
        if excel_ws==None:excel_ws=self.excel_ws
        for row in excel_ws.iter_rows(min_col=min_col, 
                                            min_row=min_row, 
                                            max_col=max_col, 
                                            max_row=max_row):
            cells.append([cell.value for cell in row])
        add_none_cells=[[None]*(max_col-min_col+2)]+[[None]+row for row in cells]
        return pd.DataFrame(add_none_cells)
    
                
    def axising_range_value_df(self,range_value_df,ws_row=None,ws_column=None,set_Nan=False):
        """获取/赋空df的特定行/列值"""
        if not set_Nan:#? 为什么不return
            if ws_row:range_value_df.iloc[ws_row,:] 
            if ws_column:range_value_df.iloc[:,ws_column]
        else:
            if ws_row:range_value_df.iloc[ws_row,:]="__SPECIAL_VALUE__"
            if ws_column:range_value_df.iloc[:,ws_column]="__SPECIAL_VALUE__"
            return range_value_df
    def get_range_cells_dict(self, min_col, min_row, max_col, max_row, value_only=True):#暂时好像不用了
        """获取指定区域内的单元格，返回一个字典，其中键是单元格的位置，值是单元格的内容或对象"""
        transform_cell = lambda cell: cell.value if value_only else cell
        cells_dict = {}
        for row in self.excel_ws.iter_rows(min_col=min_col, 
                                        min_row=min_row, 
                                        max_col=max_col, 
                                        max_row=max_row):
            for cell in row:
                cell_coordinate = cell.coordinate
                cells_dict[cell_coordinate] = transform_cell(cell)
        return cells_dict

    
    def get_max_row_col(self):
        """worksheet提供的属性来获取最大行列数问题：目前发现单元格有颜色填充、字色等也会被视为有内容的单元格；
           根据值遍历出的最大行列数则无此问题
           此外，纯下拉列表无选择值，二者都不会视为单元格有内容
           故返回两种方法分别产生的最大行列数集合"max_col""max_row"
        """
        px_max_row = self.excel_ws.max_row
        px_max_col = self.excel_ws.max_column
        value_max_row = 0
        value_max_col = 0
        last_cell=""
        for row in self.excel_ws.iter_rows():
            for cell in row:
                if cell.value:
                    if "".join(str(cell.value).split()):
                        last_cell=cell
                        value_max_row = max(value_max_row, cell.row)
                        value_max_col = max(value_max_col, cell.column)
        return {"max_col":{px_max_col,value_max_col},
                "max_row":{px_max_row,value_max_row}}
    
    def append_df_to_ws_from_row(self,df, start_row, include_index=False, include_header=False):
        """
        从指定的行开始，将pandas DataFrame添加到openpyxl worksheet中。

        :param ws: openpyxl worksheet对象。
        :param df: 要添加的pandas DataFrame对象。
        :param start_row: 开始添加的起始行。
        :param include_index: 是否包含DataFrame的索引作为额外的一列。
        :param include_header: 是否包含DataFrame的列名作为额外的一行。
        """
        ws=self.excel_ws
        # 转换DataFrame为worksheet的行
        rows = dataframe_to_rows(df, index=include_index, header=include_header)
        for r_idx, r in enumerate(rows, start=1):
            for c_idx, value in enumerate(r, start=1):
                # 从指定的start_row开始添加
                ws.cell(row=start_row + r_idx - 1, column=c_idx, value=value)
                
    def modify_cell_style(self, cell, style_dict,not_modify_attr=[]):
        """使用字典参数来修改某一单元格的样式和值属性，可自选不修改的属性
            ['font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink', 'value']"""
        # Check if cell is a string reference or a Cell object
        if isinstance(cell, str):
            cell = self.excel_ws[cell]
        # Define all possible attributes that can be modified

        possible_attributes = [
            'font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink' ,'value',"_style"]
        for key in style_dict:
            if key not in possible_attributes:
                print(f"您设置的key{key}并不属于常见的单元格属性，请自行核查")
                break
        if not_modify_attr:
            possible_attributes = [attr for attr  in possible_attributes if attr not in not_modify_attr]
        
        # copy value seperately
        if ("value" in style_dict) and ("value" not in not_modify_attr):
            self.excel_ws.cell(row=cell.row,column=cell.column,value=style_dict["value"])
        # Iterate over possible attributes and update if provided in style_dict
        for attr_name in possible_attributes:
            if attr_name in style_dict:
                setattr(cell, attr_name, copy(style_dict[attr_name]))


    def modify_CertainRange_style(self, cell_range, style_dict, not_modify_attr=[]):
        """根据字典参数修改某一单元格区域的字体、边框、填充、数字格式、保护方式、超文本、对齐格式等,
        可自选不修改的属性['font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink', 'value']"""
        # Convert cell range to actual range
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        
        # Iterate over all cells in the range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = self.excel_ws.cell(row=row, column=col)
                self.modify_cell_style(cell, style_dict, not_modify_attr = not_modify_attr )
                
    def modify_MutipleRange_style(self, cell_ranges_str, style_dict, not_modify_attr=[]):
        """根据字典参数修改某一多分布单元格区域的字体、边框、填充、数字格式、保护方式、超文本、对齐格式等，
        可自选不修改的属性['font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink', 'value']"""
        cell_ranges_list=[i.strip() for i in cell_ranges_str.split(",")]
        for cell_range in cell_ranges_list:
            self.modify_CertainRange_style(cell_range, style_dict, not_modify_attr= not_modify_attr)

    def get_cell_attributes(self, cell_reference, not_get_attr=[]):
        """获取单元格的属性字典，包括value和hyperlink,
        但可自选不修改的属性['font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink', 'value']"""
        cell = self.excel_ws[cell_reference]
        attributes = {
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': cell.number_format,
            'protection': copy(cell.protection),
            'alignment': copy(cell.alignment),
            'hyperlink': cell.hyperlink,
            'value': cell.value,
            "_style":cell._style
        }
        if len(not_get_attr)>0:
            for i in not_get_attr:
                attributes.pop(i)
        # 使用deepcopy可以确保对象是完全独立的副本
        return {cell_reference: deepcopy(attributes)}
    
    def get_row_attributes(self, row_number, not_get_attr=[]):
        """遍历获取一行所有单元格的属性字典，包括value和hyperlink,
        但可自选不修改的属性['font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink', 'value']"""
        row_attributes = {}
        for cell in self.excel_ws[row_number]:
            cell_reference = cell.coordinate
            cell_attributes = self.get_cell_attributes(cell_reference,not_get_attr=not_get_attr)
            row_attributes.update(cell_attributes)
        return row_attributes
    
    def get_CertainRange_attributes(self, cell_range, not_get_attr=[]):
        """根据字典参数遍历获取某一连续单元格区域的字体、边框、填充、数字格式、保护方式、超文本、对齐格式等，
        包括value和hyperlink,但可自选不修改的属性['font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink', 'value']"""
        cells_attributes = {}
        # Convert cell range to actual range
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        
        # Iterate over all cells in the range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                col_letter=get_column_letter(col)
                cells_attributes.update(self.get_cell_attributes(col_letter+str(row), not_get_attr=not_get_attr))
        return cells_attributes
    
    def get_MutipleRange_attributes(self, cell_ranges_str, not_get_attr=[]):
        """根据字典参数遍历获取某一多分布单元格区域的字体、边框、填充、数字格式、保护方式、超文本、对齐格式等，
        包括value和hyperlink,但可自选不修改的属性['font', 'border', 'fill', 'number_format',
            'protection', 'alignment', 'hyperlink', 'value']"""
        cells_attributes = {}
        cell_ranges_list=[i.strip() for i in cell_ranges_str.split(",")]
        for cell_range in cell_ranges_list:
            cells_attributes.update(self.get_CertainRange_attributes(cell_range, not_get_attr=not_get_attr))
        return cells_attributes
    
    def get_merge_ranges_str(self):
        merge_ranges_list = []
        for merged_cell_range in shopier.source_ws_attr.excel_ws.merged_cells.ranges:
            merge_ranges_list.append(merged_cell_range.coord)
        return merge_ranges_list
    
    
    def clear_cell_attributes(self, cell):
        """完全清空单元格的任何属性，包括hyperlink"""
        if type(cell) == str:cell=self.excel_ws[cell]
        cell.font = Font()
        cell.border = Border()
        cell.fill = PatternFill(fill_type=None)
        cell.number_format = 'General'
        cell.protection = Protection()
        cell.alignment = Alignment()
        cell.value = None
        cell.hyperlink = None
    
    

    '''def apply_cell_attributes(self, cell_reference, attributes):
        """将属性字典应用到指定单元格"""
        cell = self.excel_ws[cell_reference]
        cell.font = attributes.get('font', Font())
        cell.border = attributes.get('border', Border())
        cell.fill = attributes.get('fill', Fill())
        cell.number_format = attributes.get('number_format', 'General')
        cell.protection = attributes.get('protection', Protection())
        cell.alignment = attributes.get('alignment', Alignment())
        if 'hyperlink' in attributes and attributes['hyperlink'] is not None:
            cell.hyperlink = copy(attributes['hyperlink'])
        cell.value = attributes.get('value')'''

    # 获取特定工作表特定行的行高
    def get_row_height(self, row_number):
        return self.excel_ws.row_dimensions[row_number].height

    # 获取特定工作表特定列的列宽
    def get_column_width(self, column_number):
        return self.excel_ws.column_dimensions[get_column_letter(column_number)].width

    # 将特定工作表特定行宽设置为特定值
    def set_row_height(self, row_number, height):
        self.excel_ws.row_dimensions[row_number].height = height

    # 将特定工作表特定列宽设为特定值
    def set_column_width(self, column_number, width):
        self.excel_ws.column_dimensions[get_column_letter(column_number)].width = width
        
    def get_dropdowns(self):
        """获取工作表内的各列的下拉列表字典，同一列多种下拉列表的以list组织"""
        def get_dropdowns_values(validation):
            result=validation.formula1
            
            # 进一步，下拉列表不仅仅为序列
            # 若值为工作表单元格引用
            
            #捕获组 (.*!)? 是可选的，用来匹配任意字符后跟一个感叹号 !，代表可能存在的工作表名称。
            pattern = r"^(.*!)?(\$?[A-Za-z]\$?\d+:\$?[A-Za-z]\$?\d+)$"
            match_=re.search(pattern,result)
            if match_:
                match_groups=match_.groups()
                # 若跨工作表引用(预计更为合理)
                if (match_groups)[0]:
                    sheet_name=match_groups[0][:-1]
                    sheet_name=sheet_name[1:-1] if sheet_name[0]==sheet_name[-1]=="'" else sheet_name
                    dropdown_sourcesheet=self.excel_wb[sheet_name]
                else:dropdown_sourcesheet=self.excel_ws
                # 默认被引用为数据验证的单元格不止一个
                min_col, min_row, max_col, max_row = range_boundaries(match_groups[-1].replace('$', ''))
                value_list=[]
                for i in range(min_row, max_row+1):
                    for j in range(min_col, max_col+1):
                        value_list.append(dropdown_sourcesheet.cell(i, j).value)
                return value_list
            
            # 若值为简单的手动输入序列
            elif "," in result:
                # 去除首尾的引号后，直接拆分为值
                return result[1:-1].split(',')
            
        drop_row=dict()

        # 含有当前工作表的所有有效性验证的对象
        validations = self.excel_ws.data_validations.dataValidation
        #print(validations)
        for validation in validations:
            
            #当前有效性涉及区域
            cell=str(validation.sqref)
            
            #目前的方式，仅匹配下拉列表选择所有值的。进一步：考虑介于等多种方式
            result=(get_dropdowns_values(validation))

            #如果是多列的下拉列表相同，分别进行检验
            if " " in cell:
                cells=cell.split(" ")
                for i in cells:
                    if i[0] not in drop_row:drop_row[i[0]]=[result]
                    elif set(result) in [set(already_result) for already_result in drop_row[i[0]]]:continue
                    else:drop_row[i[0]].append(result)
            else:
                if (cell)[0] not in drop_row:drop_row[(cell)[0]]=[result]
                else:drop_row[cell[0]].append(result)
        return drop_row
         
    def create_or_update_dv_list(self, field, rule_list):
        """将过长的下拉列表中的选项写入隐藏的工作表中,工作表中各列内容：
            A           B           C
        1   A1序号      C1院系(新表列号与旧表列号不必一直)
        2   下拉列表值1  下拉列表值1  .
        3   下拉列表值2  下拉列表值2  .
        4   下拉列表值3  下拉列表值3  .
        5   下拉列表值4  下拉列表值4  .
        6   下拉列表值5  下拉列表值5  .
        7   下拉列表值6  下拉列表值6  .
        """
        # 检查是否存在名为'sheet_for_DataValidate'的工作表，如果没有则创建
        sheet_name = 'sheet_for_DataValidate'
        if sheet_name not in self.excel_wb.sheetnames:
            dv_sheet = self.excel_wb.create_sheet(sheet_name)
            dv_sheet.sheet_state = 'hidden'  # 隐藏工作表
        else:
            dv_sheet = self.excel_wb[sheet_name]

        # 查找为空的列（即没有数据验证列表的列）
        col_index = 1
        while dv_sheet.cell(row=2, column=col_index).value is not None:
            col_index += 1

        # 在找到的列的第一行中写入字段名
        dv_sheet.cell(row=1, column=col_index, value=field)

        # 从第二行开始写入规则列表
        for index, item in enumerate(rule_list, start=2):
            dv_sheet.cell(row=index, column=col_index, value=item)

        # 返回引用区域的字符串，例如'Sheet2!$B$2:$B$100'
        return f"'{sheet_name}'!${dv_sheet.cell(row=2, column=col_index).column_letter}$2:${dv_sheet.cell(row=2, column=col_index).column_letter}${len(rule_list)+1}"


    def set_dropdowns(self, selected_field_rules, sep_row=2):
        """将用户选定的规则字典导出到下拉列表，默认假设字段在第n行，在第n+1行添加规则样例行，n+2开始是下拉列表"""
        #selected_field_rules = {k: v for k, v in selected_field_rules.items() if v[1]}  # 去掉规则列表没有内容的字段
        
        for one_index_col, (field_name, rule_list) in selected_field_rules.items():
            if rule_list:  # 当规则列表有内容时
                dv_col, dv_beginrow = one_index_col[0], int(one_index_col[1:]) + sep_row
                sqref = f'{dv_col}{dv_beginrow}:{dv_col}1048576'  # 确保范围引用格式正确
                
                # 将规则列表转化为逗号分隔的字符串,并检测是否比255长，若过长即采用引用区域方式呈现。
                formula1_insides_quotes = ",".join(rule_list)
                if len(formula1_insides_quotes)>250:
                    formula1=self.create_or_update_dv_list(one_index_col+":"+field_name,rule_list)
                else:formula1=f'"{formula1_insides_quotes}"'
                # 添加下拉列表及其对应区域
                dv = DataValidation(type="list", formula1=formula1, showErrorMessage=True, allow_blank=False)
                self.excel_ws.add_data_validation(dv)
                dv.add(sqref)

    def set_validation_rules_and_example(self, one_index_col, field_name, rule_list, example):
        """#进一步：若是有限选项，就传list；若是定类，就传str pattern 屎山做法
        设置规则和样例到指定单元格，并且如果规则过长，只使用规则列表的前n项，
        使得len(",".join(rule_list[:n]))<20，并在后面加上"等{len(rule_list)}个选项"。
        同时设置单元格字体为红色。#进一步：调整样式；富文本？
        """
        # 计算合适的规则字符串长度
        rule_display = "，".join(rule_list)
        cell_value = f"·规则：{rule_display}\n·样例：{example}"
        cell = self.excel_ws[f"{one_index_col[0]}{int(one_index_col[1])+1}"]#进一步：如何确认规则样例行就在字段的下一行？请调查核实。或者说，字段行下一行如何确保没有内容？
        cell.value = cell_value
        cell.font = Font(color="FF0000",size=7)  # 红色字体
        cell.alignment = Alignment(wrapText=True)  # 开启自动换行"""

# 实现工作表之间的复制
class SheetCopier:
    def __init__(self,source_wb,target_wb):
        """先后传入源工作簿、目标工作簿,目标为None则自动创建新工作工作簿"""
        
        self.source_ws_attr=Excel_attribute(source_wb)
        self.target_ws_attr=Excel_attribute(target_wb)
    
    def copy_some_row_attr(self, 
                      source_row, target_row, 
                      not_modify_attr=[],
                      max_s_col=16384, 
                      min_s_col=1, min_t_col=1):#进一步：复制下拉列表
        """复制源工作表指定行单元格属性到目标工作表指定行，可自选不复制的属性、源起始列、源终止列、目标起始列
        属性包括['font', 'border', 'fill', 'number_format','protection', 'alignment', 'hyperlink', 'value']"""
        is_valid_col=lambda var,max_ : var > 0 and var <=max_
        
        if not (is_valid_col(min_s_col,max_s_col) and is_valid_col(min_t_col,16384) and is_valid_col(max_s_col,16384)) or min_s_col> max_s_col:
            raise
        
        # source_excel_ws=self.source_ws_attr
        source_row_attr=self.source_ws_attr.get_row_attributes(source_row)
        
        max_s_col=max(self.source_ws_attr.get_max_row_col()["max_col"]) if max_s_col==16384 else max_s_col
        s_cell_list=[f"{get_column_letter(s_col)}{source_row}" for s_col in range(min_s_col,max_s_col+1)]
        t_cell_list=[f"{StringPRO.coordinate_from_string(s_cell)[0]}{target_row}" for s_cell in s_cell_list]
        s_t_cell_map=zip(s_cell_list,t_cell_list)
        
        # Copying merged cells
        # for merged_cell_range in source_sheet.merged_cells.ranges:
        #     if merged_cell_range.min_row <= source_row <= merged_cell_range.max_row:
        #         # Reconstruct merged cell coordinates for the target
        #         target_merged_cell_range = f"{get_column_letter(merged_cell_range.min_col)}{target_row}:{get_column_letter(merged_cell_range.max_col)}{target_row}"
        #         target_sheet.merge_cells(target_merged_cell_range)
        
        for s_cell,t_cell in s_t_cell_map:
            self.target_ws_attr.modify_cell_style(t_cell,source_row_attr[s_cell],not_modify_attr=not_modify_attr)
    # 旧版
    # def copy_row_attr(self, source_row, target_row, max_col=None):
    #     """Copy cell attributes and merged cells from source row to target row."""
    #     source_sheet, target_sheet = self.source_ws_attr.excel_ws, self.target_ws_attr.excel_ws
        
    #     # Determine max_col if not provided
    #     if max_col is None:
    #         max_col = source_sheet.max_column
        
    #     # Copying merged cells
    #     for merged_cell_range in source_sheet.merged_cells.ranges:
    #         if merged_cell_range.min_row <= source_row <= merged_cell_range.max_row:
    #             # Reconstruct merged cell coordinates for the target
    #             target_merged_cell_range = f"{get_column_letter(merged_cell_range.min_col)}{target_row}:{get_column_letter(merged_cell_range.max_col)}{target_row}"
    #             target_sheet.merge_cells(target_merged_cell_range)

    #     # Copying cell attributes
    #     for column in range(1, max_col + 1):
    #         source_cell = source_sheet.cell(row=source_row, column=column)
            
    #         # First copy value
    #         target_sheet.cell(row=target_row, column=column, value= copy(source_cell.value))
    #         target_cell = target_sheet.cell(row=target_row, column=column)
            
    #         if source_cell.has_style:
    #             target_cell._style = copy(source_cell._style)
    #             target_cell.font = copy(source_cell.font)
    #             target_cell.border = copy(source_cell.border)
    #             target_cell.fill = copy(source_cell.fill)
    #             target_cell.number_format = copy(source_cell.number_format)
    #             target_cell.protection = copy(source_cell.protection)
    #             target_cell.alignment = copy(source_cell.alignment)
    #         if source_cell.hyperlink:
    #             target_cell.hyperlink = copy(source_cell.hyperlink)
    #         else:print(source_cell.value)

    # 将源文件特定列数据验证复制到目标文件特定列，从特定行开始复制，若源列有多种下拉列表，以所需下拉列表选项独具的关键词为参数找到对应下拉列表 letter -> letter
    def copy_some_col_dropdown(self,source_col,target_col,field_row,sep_row=1,dropdown_keyword=""):
        source_dropdowns=self.source_ws_attr.get_dropdowns()
        if len(source_dropdowns[source_col])>1 and len(dropdown_keyword)>0:
            choice_suits_key=0
            for dropdown_pattern in source_dropdowns[source_col]:
                for dropdown_choice in dropdown_pattern:
                    if dropdown_keyword in dropdown_choice:
                        choice_suits_key=1;break
            if choice_suits_key==0:raise KeyError
        # 未提供下拉列表关键词，或该列只含一种下拉列表，均使用第1种下拉列表
        else:
            dropdown_choice=source_dropdowns[source_col][0]
        target_dropdown={field_row:[target_col,dropdown_choice]}
        self.target_ws_attr.set_dropdowns(target_dropdown)
        
    # 将源文件特定行高复制到目标文件特定行
    def copy_some_row_height(self,source_row,target_row):
        row_height=self.source_ws_attr.get_row_height(source_row)
        self.target_ws_attr.set_row_height(target_row,row_height)
        
    # 将源工作表的某些列宽应用到目标工作表的相应列
    def copy_some_col_widths(self, start_index, end_index):
        """# 将列范围转换为列的数字表示形式
        col_start, col_end = column_range.split(':')
        start_index = column_index_from_string(col_start)
        end_index = column_index_from_string(col_end)"""

        # 遍历指定的列范围并复制宽度
        for col in range(start_index, end_index + 1):
            column_letter = get_column_letter(col)
            width = self.source_ws_attr.get_column_width(col)
            self.target_ws_attr.set_column_width(col,width)

    # 复制源工作表的所有列宽到目标工作表
    def copy_all_col_widths(self):
        self.copy_some_col_widths(1,max(self.source_ws_attr.get_max_row_col()["max_col"]))
        


# for concat：
class Df_:
    def __init__(self, file_path_list): 
        """"""
        self.data_rows={}#各表数据行数
    def compare_field_data(self, field_names, template_names): 
        """比较字段名与模板。输出"""
    def clean_data(self):
        """清洗的任务，如去除空行、修正数据格式"""
    def merge_excels(self): 
        """合并Excel文件。"""
class Style_template:
    """此类负责存储和应用样式模板。"""
    def __init__(self, workbook): 
        """初始化方法，接收一个openpyxl workbook对象。将工作表作为属性"""
    def get_style_and_location(self):
        """从模板获取样式，保存为类属性"""
    def apply_to_cell(self, cell): 
        """应用样式到单个单元格。"""
    def apply_to_header(self): 
        """应用样式到表头。"""
    def apply_to_data(self): 
        """应用样式到数据行。"""
if 1:
    def stream_files_to_zip(file_data):
        "将文件名与io.bytes的字典保存到zip流输出"

        # 创建Zip文件的io.BytesIO对象
        zip_buffer = io.BytesIO()

        # 将文件数据直接写入到zip文件
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED) as zip_file:
            for filename, data in file_data.items():
                data.seek(0)
                zip_file.writestr(filename, data.read())

        # 准备好io.BytesIO对象以供下载
        zip_buffer.seek(0)
        return zip_buffer

    def unzip_data_stream(zip_data_stream):
        """
        解压ZIP数据流，并返回内部文件的文件名与bytes数据流的字典。
        
        参数:
        - zip_data_stream: ZIP压缩包的数据流 (bytes)。
        
        返回:
        - 一个字典，键为文件名，值为对应文件的bytes数据流。
        """
        # 使用 BytesIO 将 bytes 数据流转换为一个类文件对象
        zip_in_memory = io.BytesIO(zip_data_stream)
        
        # 创建一个字典来存储文件名与bytes数据流
        file_contents = {}
        
        # 使用 zipfile 模块打开 ZIP 数据流
        with zipfile.ZipFile(zip_in_memory, 'r') as zip_ref:
            for file_name in zip_ref.namelist():
                # 读取并存储每个文件的bytes数据流
                with zip_ref.open(file_name) as file:
                    file_contents[file_name] = file.read()
                    
        return file_contents

    # 一些简单的格式转换和读取           
    def convert_to_json_stream(data):
        """将Python数据类型转化为JSON格式的字符串，后端不再使用。"""
        json_string = json.dumps(data, indent=4, ensure_ascii=False)
        
        # 创建一个StringIO对象，它提供了文件类的接口
        json_stream = io.StringIO(json_string)
        
        # 返回数据流
        return json_stream
    def save_py_objection_to_json(py_ob,path):
        with open(path, 'w', encoding='utf-8') as json_file:
            json.dump(py_ob, json_file, indent=4, ensure_ascii=False)
    def read_from_json_stream(json_stream):
        """从JSON数据流中读取数据并转换为Python数据类型，后端不再使用。"""
        # 重置流的读取位置到起始处
        json_stream.seek(0)
        
        # 从数据流中读取JSON数据并转换为Python数据类型
        data = json.load(json_stream)
        
        # 返回Python数据类型
        return data
    def read_from_json_file(file_path):
        """从JSON文件中读取数据并转换为Python数据类型"""
        with open(file_path, 'r',encoding="utf-8") as json_file:
            data = json.load(json_file)
        return data

    def clear_directory(path):
        # 检查路径是否存在
        if os.path.exists(path):
            # 遍历目录中的所有内容
            for filename in os.listdir(path):
                file_path = os.path.join(path, filename)
                try:
                    # 如果是文件夹，则递归删除
                    if os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                    # 如果是文件，则删除文件
                    else:
                        os.remove(file_path)
                except Exception as e:
                    print(f'Failed to delete {file_path}. Reason: {e}')
    

    

if "__main__" == __name__:
    # Xio=Excel_IO()
    # output=io.BytesIO()
    # excel_got=r"tests\for_xls2xlsx\xls_file.xls"
    # with open(excel_got, 'rb') as file:
    #     output.write(file.read())
    # # 重置流的位置到开始处，这样就可以从头读取
    # output.seek(0)
    # Xio.convert_excel_format(output,"xls","xlsx",True)
    # input("xls文件已转化为xlsx文件，保存在tmp目录下，请查看")

    # 进一步：字体、行宽有点不一样
    # 实验格式、值复制
    ## 创建对象
    colorful_file=r"tests\for_concat\copy_style\colorful_sheet.xlsx"
    colorful_wb=px.load_workbook(colorful_file)
    colorful_ws=colorful_wb.worksheets[0]
    new_wb=px.Workbook()
    new_ws=new_wb.worksheets[0]
    for row in range(1,10):
        for col in range(1,20):
            new_ws[f"{get_column_letter(col)}{row}"].value="我是税远志"
    shopier=SheetCopier(colorful_wb,new_wb)
    
    ## 将source的1行复制到targer的2行
    shopier.copy_some_row_attr(1, 2,["value"])
    
    ## 将source的各列宽同步到target
    shopier.copy_all_col_widths()
    
    ## 将source的1行高同步到target的2行高
    shopier.copy_some_row_height(1, 2)
    
    ## 将target的2行复制到target的5行
    # all_data_attr=shopier.source_ws_attr.get_cell_attributes(2)
    # shopier2=SheetCopier(new_wb,new_wb)
    # shopier2.copy_row_attr(2,4)#进一步，data_only=True
    
    new_wb.save(os.path.join(StringPRO.get_filepath_variables(colorful_file)["folder_path"],"copy_color_sheet.xlsx"))
    
    